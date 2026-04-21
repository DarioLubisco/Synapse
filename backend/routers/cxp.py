from fastapi import FastAPI, HTTPException, Query, Body, Request, Form, File, UploadFile
from fastapi.responses import RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List, Optional
from fastapi.staticfiles import StaticFiles
import os
import logging
import database
import csv
import io
import uuid
import shutil
import smtplib
import ssl
import base64
import pandas as pd
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from fastapi.responses import RedirectResponse, StreamingResponse
from dotenv import load_dotenv
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

load_dotenv()

GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.send']

def get_gmail_service():
    """Returns an authorized Gmail API service instance."""
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', GMAIL_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            with open('token.json', 'w') as token:
                token.write(creds.to_json())
        else:
            return None  # No token yet, must run setup_gmail.py
    return build('gmail', 'v1', credentials=creds)

logging.basicConfig(level=logging.INFO)

from fastapi import APIRouter
router = APIRouter(tags=["CxP"])



# Ensure static folder exists

import time

class PlanPagoRequest(BaseModel):
    nros_unicos: List[int]
    fecha_planificada: str
    banco: str

class ForecastEventRequest(BaseModel):
    fecha: str
    tipo_evento: str
    valor: float = 1.0

# --- Modelos de Gastos Programados ---
class ExpenseTemplateRequest(BaseModel):
    id: Optional[int] = None
    descripcion: str
    tipo: str
    monto_estimado_usd: float
    dia_mes_estimado: int

class ProgrammedExpense(BaseModel):
    id: Optional[int] = None
    descripcion: str
    tipo: str
    monto_usd: float
    fecha_proyectada: str
    estado: str = "Pendiente"

class BatchExpenseRequest(BaseModel):
    mes: int
    anio: int
    gastos: List[ProgrammedExpense]
    descripcionesAEliminar: Optional[List[str]] = None

# --- Modelos Módulo Pagos e Indexación ---
class ProveedorDescuento(BaseModel):
    DiasDesde: int
    DiasHasta: int
    Porcentaje: float
    DeduceIVA: Optional[bool] = False

class ProveedorCondicion(BaseModel):
    CodProv: str
    DiasNoIndexacion: Optional[int] = 0
    IndexaIVA: Optional[bool] = True
    BaseDiasCredito: Optional[str] = "EMISION"
    DiasVencimiento: Optional[int] = 0
    Descuentos: Optional[List[ProveedorDescuento]] = []
    DescuentoBase_Pct: Optional[float] = 0.0
    DescuentoBase_Condicion: Optional[str] = "INDEPENDIENTE"
    DescuentoBase_DeduceIVA: Optional[bool] = False
    Email: Optional[str] = None
    TipoPersona: Optional[str] = None
    DecimalesTasa: Optional[int] = 4  # 'PJ' = Juridica, 'PN' = Natural, None = auto

class InvoiceReference(BaseModel):
    CodProv: str
    NumeroD: str
    MontoRetencionBs: Optional[float] = 0.0

class DebitNoteActionRequest(BaseModel):
    Invoices: List[InvoiceReference]

class DebitNoteRegisterRequest(BaseModel):
    Invoices: List[InvoiceReference]
    NotaDebitoID: str
    ControlID: Optional[str] = None

class SettingsBulkRequest(BaseModel):
    settings: dict

class AbonoRequest(BaseModel):
    NumeroD: str
    CodProv: str
    FechaAbono: str
    MontoBsAbonado: float
    TasaCambioDiaAbono: float
    MontoUsdAbonado: float
    AplicaIndexacion: bool
    Referencia: Optional[str] = ""

@router.get("/api/settings")
async def get_settings():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Ensure Settings table exists
        cursor.execute("""
        IF NOT EXISTS (SELECT * FROM sys.objects WHERE object_id = OBJECT_ID(N'[Procurement].[Settings]') AND type in (N'U'))
        BEGIN
            CREATE TABLE [Procurement].[Settings] (
                [SettingKey] VARCHAR(50) PRIMARY KEY,
                [SettingValue] VARCHAR(255),
                [CodSucu] VARCHAR(10) DEFAULT '00000'
            );
        END
        """)
        
        cursor.execute("SELECT SettingKey, SettingValue FROM EnterpriseAdmin_AMC.Procurement.Settings WITH (NOLOCK) WHERE CodSucu = '00000'")
        results = {row[0]: row[1] for row in cursor.fetchall()}
        return results
    except Exception as e:
        logging.error(f"Error fetching settings: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.post("/api/settings/bulk")
async def update_settings_bulk(payload: SettingsBulkRequest):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        for key, value in payload.settings.items():
            cursor.execute("SELECT 1 FROM EnterpriseAdmin_AMC.Procurement.Settings WHERE SettingKey = ? AND CodSucu = '00000'", (key,))
            if cursor.fetchone():
                cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.Settings SET SettingValue = ? WHERE SettingKey = ? AND CodSucu = '00000'", (str(value), key))
            else:
                cursor.execute("INSERT INTO EnterpriseAdmin_AMC.Procurement.Settings (SettingKey, SettingValue, CodSucu) VALUES (?, ?, '00000')", (key, str(value)))
                
        conn.commit()
        return {"message": "Settings updated successfully"}
    except Exception as e:
        logging.error(f"Error updating settings: {e}", exc_info=True)
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.post("/api/plan-pagos")
async def planificar_pagos(payload: PlanPagoRequest):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Insert each record into the PagosPlanificados table
        for nro in payload.nros_unicos:
            # Check if it already exists, if so update it, otherwise insert
            check_query = "SELECT ID FROM EnterpriseAdmin_AMC.Procurement.PagosPlanificados WHERE NroUnico = ?"
            cursor.execute(check_query, (nro,))
            existing = cursor.fetchone()
            
            if existing:
                update_query = """
                    UPDATE EnterpriseAdmin_AMC.Procurement.PagosPlanificados
                    SET FechaPlanificada = ?, Banco = ?, CodUsua = 'API_USER'
                    WHERE NroUnico = ?
                """
                cursor.execute(update_query, (payload.fecha_planificada, payload.banco, nro))
            else:
                insert_query = """
                    INSERT INTO EnterpriseAdmin_AMC.Procurement.PagosPlanificados 
                    (NroUnico, FechaPlanificada, Banco, CodUsua)
                    VALUES (?, ?, ?, 'API_USER')
                """
                cursor.execute(insert_query, (nro, payload.fecha_planificada, payload.banco))
        
        conn.commit()
        return {"message": f"Successfully planned {len(payload.nros_unicos)} payment(s)."}
        
    except Exception as e:
        logging.error(f"Error saving planned payments: {e}", exc_info=True)
        if 'conn' in locals():
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@router.post("/api/plan-pagos/unplan", response_model=None)
async def unplan_pagos(payload: dict = Body(...)):
    try:
        nros_unicos = payload.get("nros_unicos", [])
        if not nros_unicos:
            return {"message": "No invoices provided to unplan."}
            
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        placeholders = ','.join(['?'] * len(nros_unicos))
        delete_query = f"DELETE FROM EnterpriseAdmin_AMC.Procurement.PagosPlanificados WHERE NroUnico IN ({placeholders})"
        cursor.execute(delete_query, nros_unicos)
        
        conn.commit()
        return {"message": f"Successfully unplanned {cursor.rowcount} payment(s)."}
    except Exception as e:
        logging.error(f"Error unplanning payments: {e}", exc_info=True)
        if 'conn' in locals():
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@router.get("/api/cuentas-por-pagar")
async def get_cuentas_por_pagar(search: str = Query("", description="Search term for NumeroD or Provider"), desde: Optional[str] = Query(None), hasta: Optional[str] = Query(None)):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Load Limit from settings
        cursor.execute("SELECT SettingValue FROM EnterpriseAdmin_AMC.Procurement.Settings WITH (NOLOCK) WHERE SettingKey = 'LimiteCarga'")
        row = cursor.fetchone()
        limit_val = int(row.SettingValue) if row and row.SettingValue and str(row.SettingValue).isdigit() else 500
        top_clause = f"TOP {limit_val}" if limit_val > 0 else "TOP 500"

        date_filter = ""
        date_params = []
        if desde:
            date_filter += " AND CAST(SAACXP.FechaE AS DATE) >= ?"
            date_params.append(desde)
        if hasta:
            date_filter += " AND CAST(SAACXP.FechaE AS DATE) <= ?"
            date_params.append(hasta)
            
        if not desde and not hasta:
            date_filter = " AND SAACXP.FechaE >= DATEADD(month, -4, GETDATE())"
            
        # Load Tolerance from settings
        cursor.execute("SELECT SettingValue FROM EnterpriseAdmin_AMC.Procurement.Settings WITH (NOLOCK) WHERE SettingKey = 'ToleranceSaldo'")
        t_row = cursor.fetchone()
        tolerance = float(t_row.SettingValue) if t_row and t_row.SettingValue else 0.50

        # Cleanup: Remove from planning any invoice that has been fully paid (Saldo near 0)
        cleanup_query = """
            DELETE PP
            FROM EnterpriseAdmin_AMC.Procurement.PagosPlanificados PP
            INNER JOIN EnterpriseAdmin_AMC.dbo.SAACXP ON PP.NroUnico = SAACXP.NroUnico
            WHERE SAACXP.Saldo <= ? AND SAACXP.TipoCxP = '10'
        """
        cursor.execute(cleanup_query, (tolerance,))
        conn.commit()
        
        # Build query
        query = f"""
            SELECT {top_clause}
              SACOMP.FechaI,
              SACOMP.FechaE,
              SACOMP.FechaV,
              SAPROV.Descrip,
              SAACXP.RetenIVA,
              SAACXP.SaldoAct,
              SAACXP.Monto,
              SAACXP.CodOper,
              SAACXP.MontoNeto,
              CASE WHEN SAACXP.Saldo <= 0 OR SAACXP.CancelC >= SAACXP.Monto THEN 0 ELSE (ISNULL(SACOMP.MtoTotal, SAACXP.Monto) - ISNULL(SACOMP.Contado, 0) - ISNULL(SACOMP.MtoPagos, 0)) END AS Saldo,
              SAACXP.MtoTax,
              SACOMP.MtoPagos,
              SACOMP.SaldoAct AS SaldoAct_SACOMP,
              SACOMP.MtoNCredito,
              SACOMP.MtoNDebito,
              SACOMP.Signo,
              SACOMP.NumeroD AS NumeroD_SACOMP,
              SAACXP.NroCtrol,
              SACOMP.MtoTotal,
              SACOMP.Contado,
              SACOMP.Credito,
              SAACXP.NroUnico,
              SAACXP.CodSucu,
              SAACXP.CodProv,
              SAACXP.NumeroD,
              SACOMP.CodSucu AS CodSucu_SACOMP,
              SACOMP.TipoCom,
              SACOMP.Notas10,
              SACOMP.TGravable,
              SACOMP.Factor,
              SACOMP.MontoMEx,
              SACOMP.TotalPrd,
              SACOMP.Descto1,
              SACOMP.Descto2,
              SACOMP.Fletes,
              inv_set.AplicaIndexacion AS AplicaIndexacionOverride,
              SAPAGCXP.NumeroD AS NumeroD_SAPAGCXP,
              CASE WHEN ISNULL(SAACXP.Factor, 0) > 1 THEN SAACXP.Factor ELSE dt_emision.dolarbcv END AS TasaEmision,
              dt_actual.dolarbcv AS TasaActual,
              PP.ID AS Plan_ID,
              PP.Banco AS Plan_Banco,
              PP.FechaPlanificada AS Plan_Fecha,
              CAST(CASE WHEN SAACXP.RetenIVA > 0 OR portal_ret.Id IS NOT NULL THEN 1 ELSE 0 END AS BIT) AS Has_Retencion,
              CAST(CASE WHEN abonos.TotalBs IS NOT NULL THEN 1 ELSE 0 END AS BIT) AS Has_Abonos,
              ISNULL(abonos.TotalBs, 0) AS TotalBsAbonado,
              ISNULL(abonos.TotalUsd, 0) AS TotalUsdAbonado,
              ISNULL(CASE WHEN abonos.TotalIVA > 0 THEN abonos.TotalIVA ELSE portal_ret.MontoRetenido END, 0) AS RetencionIvaAbonada,
              ISNULL(abonos.TotalISLR, 0) AS RetencionIslrAbonada,
              ISNULL(SAPROV.PorctRet, 0) AS PorctRet,
              ISNULL(SAPROV.EsReten, 0) AS EsReten,
              SAPROV.ID3 AS RIF,
              SAPROV.Descrip AS ProveedorNombre
            FROM dbo.SAACXP
            OUTER APPLY (
                SELECT SUM(MontoBsAbonado) AS TotalBs,
                       SUM(CASE WHEN MontoUsdAbonado IS NOT NULL AND MontoUsdAbonado > 0 THEN MontoUsdAbonado ELSE ROUND((MontoBsAbonado / NULLIF(TasaCambioDiaAbono, 0)), 2) END) AS TotalUsd,
                       SUM(CASE WHEN TipoAbono = 'RETENCION_IVA' THEN MontoBsAbonado ELSE 0 END) AS TotalIVA,
                       SUM(CASE WHEN TipoAbono = 'RETENCION_ISLR' THEN MontoBsAbonado ELSE 0 END) AS TotalISLR
                FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos A 
                WHERE A.CodProv = SAACXP.CodProv AND A.NumeroD = SAACXP.NumeroD AND ISNULL(A.AfectaSaldo, 1) = 1
            ) abonos
            OUTER APPLY (
                SELECT TOP 1 Id, MontoRetenido
                FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA WITH (NOLOCK)
                WHERE CodProv = SAACXP.CodProv AND NumeroD = SAACXP.NumeroD AND Estado != 'ANULADO'
            ) portal_ret
            OUTER APPLY (
                SELECT TOP 1 NumeroD
                FROM dbo.SAPAGCXP
                WHERE SAPAGCXP.NroUnico = SAACXP.NroUnico
            ) SAPAGCXP
            LEFT OUTER JOIN dbo.SAPROV ON SAACXP.CodProv = SAPROV.CodProv
            LEFT OUTER JOIN dbo.SAIPACXP ON SAACXP.NroUnico = SAIPACXP.NroUnico
            LEFT OUTER JOIN dbo.SACOMP ON SAACXP.NumeroD = SACOMP.NumeroD AND SAACXP.CodProv = SACOMP.CodProv
            LEFT OUTER JOIN EnterpriseAdmin_AMC.Procurement.InvoiceSettings inv_set ON SAACXP.CodProv = inv_set.CodProv AND SAACXP.NumeroD = inv_set.NumeroD
            OUTER APPLY (
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday 
                WHERE CAST(fecha AS DATE) <= CAST(SAACXP.FechaE AS DATE)
                ORDER BY fecha DESC
            ) dt_emision
            OUTER APPLY (
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday 
                WHERE dolarbcv IS NOT NULL
                ORDER BY id DESC
            ) dt_actual
            LEFT OUTER JOIN EnterpriseAdmin_AMC.Procurement.PagosPlanificados PP
                ON SAACXP.NroUnico = PP.NroUnico
            WHERE SAACXP.TipoCxP = '10' 
               AND (SAACXP.NumeroD LIKE ?
               OR SACOMP.NumeroD LIKE ?
               OR SAPAGCXP.NumeroD LIKE ?
               OR SAPROV.Descrip LIKE ?)
               {date_filter}
            ORDER BY SAACXP.FechaE DESC
        """
        
        search_param = f"%{search}%"
        params = [search_param, search_param, search_param, search_param] + date_params
            
        cursor.execute(query, tuple(params))
        
        columns = [column[0] for column in cursor.description]
        results = []
        for row in cursor.fetchall():
            d = dict(zip(columns, row))
            for k, v in d.items():
                if hasattr(v, 'quantize') or hasattr(v, 'as_tuple'):
                    d[k] = float(v) if v is not None else 0.0
            results.append(d)
            
        return {"data": results}
        
    except Exception as e:
        logging.error(f"Error fetching data: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

# --- REPORTS ENDPOINTS ---

@router.get("/api/reports/compras")
async def report_compras(desde: Optional[str] = Query(None), hasta: Optional[str] = Query(None)):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        date_filter = ""
        date_params = []
        if desde:
            date_filter += " AND CAST(SACOMP.FechaE AS DATE) >= ?"
            date_params.append(desde)
        if hasta:
            date_filter += " AND CAST(SACOMP.FechaE AS DATE) <= ?"
            date_params.append(hasta)
            
        if not desde and not hasta:
            date_filter = " AND SACOMP.FechaE >= DATE_SUB(CURDATE(), INTERVAL 1 YEAR)" if "MySQL" in database.DRIVER else " AND SACOMP.FechaE >= DATEADD(year, -1, GETDATE())"

        # Query for grouped stats in USD
        query = f"""
            SELECT
              SACOMP.Descrip AS Proveedor,
              SUM(SACOMP.MtoTotal / NULLIF(dt_emision.dolarbcv, 0)) AS TotalUSD,
              COUNT(SACOMP.NroUnico) AS CantidadFacturas
            FROM dbo.SACOMP WITH (NOLOCK)
            OUTER APPLY (
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday WITH (NOLOCK)
                WHERE CAST(fecha AS DATE) <= CAST(SACOMP.FechaE AS DATE)
                ORDER BY fecha DESC
            ) dt_emision
            WHERE 1=1 {date_filter}
            GROUP BY SACOMP.Descrip
            ORDER BY TotalUSD DESC
        """
        cursor.execute(query, tuple(date_params))
        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        data = [dict(zip(columns, row)) for row in rows]
        
        # Calculate Grand Total for percentages
        grand_total = sum(item['TotalUSD'] for item in data if item['TotalUSD'])
        for item in data:
            item['TotalUSD'] = float(item['TotalUSD']) if item['TotalUSD'] else 0
            item['Porcentaje'] = (item['TotalUSD'] / grand_total * 100) if grand_total > 0 else 0
            
        return {"data": data, "grand_total": float(grand_total) if grand_total else 0}
    except Exception as e:
        logging.error(f"Error in report_compras: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

# --- ENDPOINTS PROVEEDORES INDEXACION ---
@router.get("/api/procurement/providers")
async def get_provider_conditions():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Return all active providers from SAPROV and their conditions if any
        query = """
            SELECT p.CodProv, p.Descrip, p.activo, p.DiasCred AS SaprovDiasCred,
                   ISNULL(c.DiasNoIndexacion, 0) AS DiasNoIndexacion, 
                   ISNULL(c.BaseDiasCredito, 'EMISION') AS BaseDiasCredito, 
                   ISNULL(c.DiasVencimiento, p.DiasCred) AS DiasVencimiento,
                   ISNULL(c.IndexaIVA, 1) AS IndexaIVA,
                   ISNULL(c.DecimalesTasa, 4) AS DecimalesTasa,
                   ISNULL(c.DescuentoBase_Pct, 0) AS DescuentoBase_Pct,
                   ISNULL(c.DescuentoBase_Condicion, 'INDEPENDIENTE') AS DescuentoBase_Condicion,
                   ISNULL(c.DescuentoBase_DeduceIVA, 0) AS DescuentoBase_DeduceIVA,
                   COALESCE(c.Email, p.Email, '') AS Email,
                   -- TipoPersona: LOCAL overrides SAINT derivation
                   COALESCE(
                       c.TipoPersona,
                       CASE WHEN LEFT(LTRIM(ISNULL(p.ID3,'')),1) IN ('J','G','C') THEN 'PJ'
                            WHEN LEFT(LTRIM(ISNULL(p.ID3,'')),1) IN ('V','E','P') THEN 'PN'
                            ELSE NULL END
                   ) AS TipoPersona,
                   c.TipoPersona AS TipoPersonaLocal,  -- explicit local override for display
                   p.ID3 AS RIF
            FROM EnterpriseAdmin_AMC.dbo.SAPROV p WITH (NOLOCK)
            LEFT JOIN EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones c WITH (NOLOCK) ON p.CodProv = c.CodProv
            ORDER BY p.Descrip
        """
        cursor.execute(query)
        columns = [column[0] for column in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Fetch dynamic discounts
        cursor.execute("SELECT CodProv, DiasDesde, DiasHasta, Porcentaje, DeduceIVA FROM EnterpriseAdmin_AMC.Procurement.ProveedorDescuentosProntoPago WITH (NOLOCK)")
        desc_cols = [c[0] for c in cursor.description]
        descuentos_db = [dict(zip(desc_cols, row)) for row in cursor.fetchall()]
        
        desc_map = {}
        for d in descuentos_db:
            desc_map.setdefault(d['CodProv'], []).append({
                "DiasDesde": int(d['DiasDesde']),
                "DiasHasta": int(d['DiasHasta']),
                "Porcentaje": float(d['Porcentaje']),
                "DeduceIVA": bool(d.get('DeduceIVA', False))
            })
            
        for r in results:
            r['Descuentos'] = desc_map.get(r['CodProv'], [])

            
        return {"data": results}
    except Exception as e:
        logging.error(f"Error loading providers: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@router.put("/api/procurement/providers/{cod_prov:path}")
async def update_provider_condition(cod_prov: str, payload: ProveedorCondicion):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Ensure empty string is saved as NULL so COALESCE fallback works
        tipo_pers = payload.TipoPersona if payload.TipoPersona and payload.TipoPersona.strip() else None
        
        # Check if condition info exists
        check_query = "SELECT CodProv FROM EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones WHERE CodProv = ?"
        cursor.execute(check_query, (cod_prov,))
        if cursor.fetchone():
            update_query = """
                UPDATE EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones
                SET DiasNoIndexacion = ?, BaseDiasCredito = ?, DiasVencimiento = ?,
                    Email = ?, TipoPersona = ?, IndexaIVA = ?, DecimalesTasa = ?,
                    DescuentoBase_Pct = ?, DescuentoBase_Condicion = ?, DescuentoBase_DeduceIVA = ?
                WHERE CodProv = ?
            """
            cursor.execute(update_query, (
                payload.DiasNoIndexacion, payload.BaseDiasCredito, payload.DiasVencimiento,
                payload.Email, tipo_pers, payload.IndexaIVA, payload.DecimalesTasa,
                payload.DescuentoBase_Pct, payload.DescuentoBase_Condicion, payload.DescuentoBase_DeduceIVA,
                cod_prov
            ))
        else:
            insert_query = """
                INSERT INTO EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones 
                (CodProv, DiasNoIndexacion, BaseDiasCredito, DiasVencimiento, Email, TipoPersona, IndexaIVA, DecimalesTasa, DescuentoBase_Pct, DescuentoBase_Condicion, DescuentoBase_DeduceIVA)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            cursor.execute(insert_query, (
                cod_prov, payload.DiasNoIndexacion, payload.BaseDiasCredito, payload.DiasVencimiento,
                payload.Email, tipo_pers, payload.IndexaIVA, payload.DecimalesTasa, payload.DescuentoBase_Pct, payload.DescuentoBase_Condicion, payload.DescuentoBase_DeduceIVA
            ))
            
        # Update dynamic discounts
        cursor.execute("DELETE FROM EnterpriseAdmin_AMC.Procurement.ProveedorDescuentosProntoPago WHERE CodProv = ?", (cod_prov,))
        if payload.Descuentos:
            for desc in payload.Descuentos:
                cursor.execute(
                    "INSERT INTO EnterpriseAdmin_AMC.Procurement.ProveedorDescuentosProntoPago (CodProv, DiasDesde, DiasHasta, Porcentaje, DeduceIVA) VALUES (?, ?, ?, ?, ?)",
                    (cod_prov, desc.DiasDesde, desc.DiasHasta, desc.Porcentaje, desc.DeduceIVA)
                )

            
        # Synchronize SAPROV native credit days
        update_saprov = "UPDATE EnterpriseAdmin_AMC.dbo.SAPROV SET DiasCred = ? WHERE CodProv = ?"
        cursor.execute(update_saprov, (payload.DiasVencimiento, cod_prov))
        
        conn.commit()
        return {"message": "Condiciones del proveedor actualizadas."}
    except Exception as e:
        logging.error(f"Error updating provider {cod_prov}: {e}", exc_info=True)
        if 'conn' in locals():
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

def enviar_correo_pago(destinatario: str, proveedor_nombre: str, nro_factura: str, pago_data, filepaths=None):
    """Send payment notification via Gmail API.
    pago_data can be a single dict or a list of dicts (for multi-invoice payments).
    Supports multiple recipients separated by ';' in destinatario.
    """
    try:
        service = get_gmail_service()
        if not service:
            logging.warning("Gmail API no disponible. Ejecuta: .venv\\Scripts\\python.exe setup_gmail.py")
            return False
        
        emails = [e.strip() for e in destinatario.split(";") if e.strip()]
        if not emails:
            logging.warning(f"No valid email addresses found in: {destinatario}")
            return False
        
        # Normalize pago_data to list
        if isinstance(pago_data, dict):
            pagos_list = [pago_data]
        else:
            pagos_list = list(pago_data)
        
        is_multi = len(pagos_list) > 1
        
        remitente = os.getenv("SMTP_EMAIL", "")
        msg = MIMEMultipart()
        msg['From'] = remitente
        msg['To'] = ", ".join(emails)
        
        if is_multi:
            total_bs = sum(float(p.get('MontoBsAbonado', 0)) for p in pagos_list)
            msg['Subject'] = f"Soporte de Pago - {len(pagos_list)} Facturas - {proveedor_nombre}"
            facturas_str = ", ".join(p.get('NumeroD', '?') for p in pagos_list)
            cuerpo = f"""Estimados/as {proveedor_nombre},

Adjunto a este correo se encuentra el resumen de pagos correspondiente a las siguientes facturas: {facturas_str}.
El monto total pagado es de: Bs. {total_bs:,.2f}

Atentamente,
El equipo de Administracion."""
        else:
            monto_bs = pagos_list[0].get('MontoBsAbonado', 0)
            msg['Subject'] = f"Soporte de Pago y Resumen - Factura {nro_factura}"
            cuerpo = f"""Estimados/as {proveedor_nombre},

Adjunto a este correo se encuentra el soporte de pago correspondiente a la factura Nro: {nro_factura}.
El monto pagado total es de: Bs. {monto_bs:,.2f}

Atentamente,
El equipo de Administracion."""
        msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
        
        # Generar Excel con formato profesional usando openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Resumen Pago'
        
        headers = ['Nro Factura', 'Referencia', 'Fecha de Pago', 'Monto Pagado (Bs)', 'Tasa de Cambio', 'Indexado', 'Monto Pagado (USD)']
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        data_font = Font(name='Calibri', size=11)
        money_fmt = '#,##0.00'
        rate_fmt = '#,##0.0000'
        
        # Fetch retentions to include them in the excel balance
        rets_iva = {}
        rets_islr = {}
        try:
            import database
            conn_ret_xl = database.get_db_connection()
            csr_xl = conn_ret_xl.cursor()
            numeros_d_xl = [p.get('NumeroD') for p in pagos_list if p.get('NumeroD')]
            cod_prov_xl = pagos_list[0].get('CodProv') if pagos_list else None
            
            if numeros_d_xl and cod_prov_xl:
                placeholders = ','.join(['?'] * len(numeros_d_xl))
                csr_xl.execute(f"SELECT NumeroD, SUM(ISNULL(MontoRetenido, 0)) FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA WHERE CodProv = ? AND NumeroD IN ({placeholders}) AND Estado <> 'ANULADO' GROUP BY NumeroD", [cod_prov_xl] + numeros_d_xl)
                for r_ in csr_xl.fetchall(): rets_iva[r_[0]] = float(r_[1])
                
                csr_xl.execute(f"SELECT NumeroD, SUM(ISNULL(MontoRetenido, 0)) FROM EnterpriseAdmin_AMC.Procurement.Retenciones_ISLR WHERE CodProv = ? AND NumeroD IN ({placeholders}) AND Estado <> 'ANULADO' GROUP BY NumeroD", [cod_prov_xl] + numeros_d_xl)
                for r_ in csr_xl.fetchall(): rets_islr[r_[0]] = float(r_[1])
        except Exception as e:
            logging.error(f"Error cargando retenciones excel: {e}")
        finally:
            if 'conn_ret_xl' in locals(): conn_ret_xl.close()

        row_idx = 2
        sum_bs = 0.0
        sum_usd = 0.0
        
        for p in pagos_list:
            ws.cell(row=row_idx, column=1, value=p.get('NumeroD', '')).font = data_font
            ws.cell(row=row_idx, column=2, value=p.get('Referencia', '')).font = data_font
            ws.cell(row=row_idx, column=3, value=p.get('FechaAbono', '')).font = data_font
            
            monto_bs = float(p.get('MontoBsAbonado', 0))
            tasa = float(p.get('TasaCambioDiaAbono', 0))
            monto_usd = float(p.get('MontoUsdAbonado', 0))
            
            sum_bs += monto_bs
            sum_usd += monto_usd
            
            c_bs = ws.cell(row=row_idx, column=4, value=monto_bs)
            c_bs.font = data_font
            c_bs.number_format = money_fmt
            c_bs.alignment = Alignment(horizontal='right')
            
            c_tasa = ws.cell(row=row_idx, column=5, value=tasa)
            c_tasa.font = data_font
            c_tasa.number_format = rate_fmt
            c_tasa.alignment = Alignment(horizontal='right')
            
            ws.cell(row=row_idx, column=6, value=p.get('AplicaIndexacion', 'No')).font = data_font
            
            c_usd = ws.cell(row=row_idx, column=7, value=monto_usd)
            c_usd.font = data_font
            c_usd.number_format = money_fmt
            c_usd.alignment = Alignment(horizontal='right')
            
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).border = thin_border
            
            row_idx += 1
            
            # Print Retencion IVA
            num_d = p.get('NumeroD', '')
            if num_d in rets_iva and rets_iva[num_d] > 0:
                ret_bs = rets_iva[num_d]
                ret_usd = (ret_bs / tasa) if tasa > 0 else 0.0
                sum_bs += ret_bs
                sum_usd += ret_usd
                
                ws.cell(row=row_idx, column=1, value=num_d).font = data_font
                ws.cell(row=row_idx, column=2, value="Retención IVA").font = data_font
                ws.cell(row=row_idx, column=3, value=p.get('FechaAbono', '')).font = data_font
                
                c_rbs = ws.cell(row=row_idx, column=4, value=ret_bs)
                c_rbs.font = data_font; c_rbs.number_format = money_fmt; c_rbs.alignment = Alignment(horizontal='right')
                
                c_rtasa = ws.cell(row=row_idx, column=5, value=tasa)
                c_rtasa.font = data_font; c_rtasa.number_format = rate_fmt; c_rtasa.alignment = Alignment(horizontal='right')
                
                ws.cell(row=row_idx, column=6, value="-").font = data_font
                
                c_rusd = ws.cell(row=row_idx, column=7, value=ret_usd)
                c_rusd.font = data_font; c_rusd.number_format = money_fmt; c_rusd.alignment = Alignment(horizontal='right')
                
                for col in range(1, 8):
                    ws.cell(row=row_idx, column=col).border = thin_border
                
                row_idx += 1
                
            # Print Retencion ISLR
            if num_d in rets_islr and rets_islr[num_d] > 0:
                ret_bs = rets_islr[num_d]
                ret_usd = (ret_bs / tasa) if tasa > 0 else 0.0
                sum_bs += ret_bs
                sum_usd += ret_usd
                
                ws.cell(row=row_idx, column=1, value=num_d).font = data_font
                ws.cell(row=row_idx, column=2, value="Retención ISLR").font = data_font
                ws.cell(row=row_idx, column=3, value=p.get('FechaAbono', '')).font = data_font
                
                c_rbs = ws.cell(row=row_idx, column=4, value=ret_bs)
                c_rbs.font = data_font; c_rbs.number_format = money_fmt; c_rbs.alignment = Alignment(horizontal='right')
                
                c_rtasa = ws.cell(row=row_idx, column=5, value=tasa)
                c_rtasa.font = data_font; c_rtasa.number_format = rate_fmt; c_rtasa.alignment = Alignment(horizontal='right')
                
                ws.cell(row=row_idx, column=6, value="-").font = data_font
                
                c_rusd = ws.cell(row=row_idx, column=7, value=ret_usd)
                c_rusd.font = data_font; c_rusd.number_format = money_fmt; c_rusd.alignment = Alignment(horizontal='right')
                
                for col in range(1, 8):
                    ws.cell(row=row_idx, column=col).border = thin_border
                
                row_idx += 1
        
        # Totals row
        if is_multi or rets_iva or rets_islr:
            tot_row = row_idx
            total_font = Font(name='Calibri', bold=True, size=11)
            total_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
            
            ws.cell(row=tot_row, column=1, value='TOTALES').font = total_font
            ws.cell(row=tot_row, column=1).fill = total_fill
            for col in range(2, 4):
                ws.cell(row=tot_row, column=col).fill = total_fill
            
            c_total_bs = ws.cell(row=tot_row, column=4, value=sum_bs)
            c_total_bs.font = total_font
            c_total_bs.fill = total_fill
            c_total_bs.number_format = money_fmt
            c_total_bs.alignment = Alignment(horizontal='right')
            
            ws.cell(row=tot_row, column=5).fill = total_fill
            ws.cell(row=tot_row, column=6).fill = total_fill
            
            c_total_usd = ws.cell(row=tot_row, column=7, value=sum_usd)
            c_total_usd.font = total_font
            c_total_usd.fill = total_fill
            c_total_usd.number_format = money_fmt
            c_total_usd.alignment = Alignment(horizontal='right')
            
            for col in range(1, 8):
                ws.cell(row=tot_row, column=col).border = thin_border
        
        # Auto-fit column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(max_length + 3, 12)
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        part_excel = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part_excel.set_payload(excel_buffer.read())
        encoders.encode_base64(part_excel)
        fn_suffix = f"{len(pagos_list)}_Facturas" if is_multi else nro_factura
        part_excel.add_header("Content-Disposition", f"attachment; filename=Resumen_Pago_{fn_suffix}.xlsx")
        msg.attach(part_excel)
        
        import json as _json
        all_att_paths = set()
        
        if isinstance(filepaths, list):
            for fp in filepaths:
                if fp: all_att_paths.add(fp)
        elif filepaths:
            all_att_paths.add(filepaths)
            
        numeros_d = [p.get('NumeroD') for p in pagos_list if p.get('NumeroD')]
        cod_prov = pagos_list[0].get('CodProv') if pagos_list else None
        
        try:
            if numeros_d and cod_prov:
                import database
                conn_hist = database.get_db_connection()
                cursor_hist = conn_hist.cursor()
                placeholders = ','.join(['?'] * len(numeros_d))
                cursor_hist.execute(f"SELECT RutaComprobante FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos WHERE CodProv = ? AND NumeroD IN ({placeholders}) AND RutaComprobante IS NOT NULL AND RutaComprobante <> ''", [cod_prov] + numeros_d)
                for r in cursor_hist.fetchall():
                    ruta = r[0]
                    if ruta:
                        ruta = ruta.strip()
                        if ruta.startswith('['):
                            try:
                                paths = _json.loads(ruta)
                                for p_ in paths:
                                    if p_: all_att_paths.add(p_)
                            except: pass
                        else:
                            all_att_paths.add(ruta)
        except Exception as e:
            logging.error(f"Error extraeyendo historial de comprobantes: {e}", exc_info=True)
        finally:
            if 'conn_hist' in locals(): conn_hist.close()
            
        for idx_att, fp in enumerate(all_att_paths):
            if fp and os.path.exists(fp):
                with open(fp, "rb") as attachment:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(attachment.read())
                encoders.encode_base64(part)
                ext = os.path.splitext(fp)[1]
                part.add_header("Content-Disposition", f"attachment; filename=Soporte_{idx_att+1}_{fn_suffix}{ext}")
                msg.attach(part)
        
        # ==========================================
        # ANEXAR RETENCIONES DE IVA E ISLR (SI EXISTEN y ESTAN GENERADAS)
        # ==========================================
        try:
            import database
            conn_ret = database.get_db_connection()
            cursor_ret = conn_ret.cursor()
            
            if numeros_d and cod_prov:
                placeholders = ','.join(['?'] * len(numeros_d))
                
                # Obtener configuración para el PDF
                cursor_ret.execute("SELECT RIF_Agente, RazonSocial_Agente, DireccionFiscal_Agente, ValorUT, UltimoSecuencial FROM EnterpriseAdmin_AMC.Procurement.Retenciones_Config WHERE Id = 1")
                cfg_row = cursor_ret.fetchone()
                config = dict(zip([c[0] for c in cursor_ret.description], cfg_row)) if cfg_row else {}
                
                from itertools import groupby
                
                # --- IVA ---
                cursor_ret.execute(f"SELECT r.*, p.Descrip as ProveedorNombre FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA r LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON r.CodProv = p.CodProv WHERE r.CodProv = ? AND r.NumeroD IN ({placeholders}) AND r.Estado <> 'ANULADO'", [cod_prov] + numeros_d)
                iva_rows = cursor_ret.fetchall()
                if iva_rows:
                    cols = [c[0] for c in cursor_ret.description]
                    iva_list = [dict(zip(cols, row)) for row in iva_rows]
                    iva_list.sort(key=lambda x: str(x.get('NumeroComprobante', '')))
                    for nro_comp, group in groupby(iva_list, key=lambda x: str(x.get('NumeroComprobante', ''))):
                        if not nro_comp or nro_comp == 'None': continue
                        g_list = list(group)
                        # calls generar_pdf_retencion defined later
                        pdf_bytes = generar_pdf_retencion(config, g_list)
                        part_pdf = MIMEBase("application", "pdf")
                        part_pdf.set_payload(pdf_bytes)
                        encoders.encode_base64(part_pdf)
                        part_pdf.add_header("Content-Disposition", f"attachment; filename=Retencion_IVA_{nro_comp}.pdf")
                        msg.attach(part_pdf)

                # --- ISLR ---
                cursor_ret.execute(f"SELECT r.*, p.Descrip as ProveedorNombre FROM EnterpriseAdmin_AMC.Procurement.Retenciones_ISLR r LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON r.CodProv = p.CodProv WHERE r.CodProv = ? AND r.NumeroD IN ({placeholders}) AND r.Estado <> 'ANULADO'", [cod_prov] + numeros_d)
                islr_rows = cursor_ret.fetchall()
                if islr_rows:
                    cols = [c[0] for c in cursor_ret.description]
                    islr_list = [dict(zip(cols, row)) for row in islr_rows]
                    islr_list.sort(key=lambda x: str(x.get('NumeroComprobante', '')))
                    for nro_comp, group in groupby(islr_list, key=lambda x: str(x.get('NumeroComprobante', ''))):
                        if not nro_comp or nro_comp == 'None': continue
                        g_list = list(group)
                        try:
                            pdf_bytes = generar_pdf_islr(config, g_list)
                            part_pdf = MIMEBase("application", "pdf")
                            part_pdf.set_payload(pdf_bytes)
                            encoders.encode_base64(part_pdf)
                            part_pdf.add_header("Content-Disposition", f"attachment; filename=Retencion_ISLR_{nro_comp}.pdf")
                            msg.attach(part_pdf)
                        except NameError:
                            logging.warning("generar_pdf_islr no definido, no se anexa PDF de ISLR.")
                        except Exception as e:
                            logging.error(f"Error generando PDF ISLR: {e}", exc_info=True)
                            
        except Exception as e:
            logging.error(f"Error adjuntando retenciones: {e}", exc_info=True)
        finally:
            if 'conn_ret' in locals(): conn_ret.close()
        
        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        service.users().messages().send(userId='me', body={'raw': raw}).execute()
        logging.info(f"Correo enviado exitosamente a {', '.join(emails)} via Gmail API")
        return True
    except HttpError as e:
        logging.error(f"Gmail API HttpError al enviar correo a {destinatario}: {e}")
        return False
    except Exception as e:
        logging.error(f"Error al enviar correo a {destinatario}: {e}")
        return False

# Helper: wrap email send so a network error doesn't crash the payment transaction
def safe_send_email(destinatario: str, proveedor_nombre: str, nro_factura: str, pago_data, filepaths=None) -> bool:
    try:
        return enviar_correo_pago(destinatario, proveedor_nombre, nro_factura, pago_data, filepaths)
    except Exception as e:
        logging.warning(f"Email send failed (likely offline): {e}")
        return False


# ==============================================================================
# MOTIVOS DE AJUSTE - CRUD
# ==============================================================================
@router.get("/api/procurement/motivos-ajuste")
async def get_motivos_ajuste(solo_activos: bool = True):
    """Lista todos los motivos de ajuste contable."""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        query = "SELECT MotivoID, Codigo, Descripcion, AnulaNotaDebito, AnulaNotaCredito, ParaAjuste, ParaNotaCredito, Activo, EmailTemplate FROM EnterpriseAdmin_AMC.Procurement.MotivosAjuste"
        if solo_activos:
            query += " WHERE Activo = 1"
        query += " ORDER BY Codigo"
        cursor.execute(query)
        columns = [c[0] for c in cursor.description]
        rows = [dict(zip(columns, r)) for r in cursor.fetchall()]
        for r in rows:
            r['AnulaNotaDebito']  = bool(r['AnulaNotaDebito'])
            r['AnulaNotaCredito'] = bool(r['AnulaNotaCredito'])
            r['ParaAjuste']       = bool(r['ParaAjuste'])
            r['ParaNotaCredito']  = bool(r['ParaNotaCredito'])
            r['Activo']           = bool(r['Activo'])
        return {"data": rows}
    except Exception as e:
        logging.error(f"Error get_motivos_ajuste: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.post("/api/procurement/motivos-ajuste")
async def upsert_motivo_ajuste(payload: dict = Body(...)):
    """Crea o actualiza un motivo de ajuste contable."""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        motivo_id = payload.get("MotivoID")
        email_template = payload.get("EmailTemplate", "")
        codigo      = payload.get("Codigo", "")
        descripcion = payload.get("Descripcion", "")
        anula_nd    = 1 if payload.get("AnulaNotaDebito",  False) else 0
        anula_nc    = 1 if payload.get("AnulaNotaCredito", False) else 0
        para_ajuste = 1 if payload.get("ParaAjuste", True) else 0
        para_nc     = 1 if payload.get("ParaNotaCredito", False) else 0
        activo      = 1 if payload.get("Activo", True)           else 0

        if motivo_id:
            cursor.execute("""
                UPDATE EnterpriseAdmin_AMC.Procurement.MotivosAjuste
                SET Codigo = ?, Descripcion = ?, AnulaNotaDebito = ?, AnulaNotaCredito = ?, ParaAjuste = ?, ParaNotaCredito = ?, Activo = ?, EmailTemplate = ?
                WHERE MotivoID = ?
            """, (codigo, descripcion, anula_nd, anula_nc, para_ajuste, para_nc, activo, email_template, motivo_id))
            msg = "Motivo actualizado."
        else:
            cursor.execute("""
                INSERT INTO EnterpriseAdmin_AMC.Procurement.MotivosAjuste
                    (Codigo, Descripcion, AnulaNotaDebito, AnulaNotaCredito, ParaAjuste, ParaNotaCredito, Activo, EmailTemplate)
                OUTPUT INSERTED.MotivoID
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (codigo, descripcion, anula_nd, anula_nc, para_ajuste, para_nc, activo, email_template))
            motivo_id = cursor.fetchone()[0]
            msg = "Motivo creado."

        conn.commit()
        return {"data": {"MotivoID": motivo_id}, "message": msg}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error upsert_motivo_ajuste: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.delete("/api/procurement/motivos-ajuste/{motivo_id}")
async def delete_motivo_ajuste(motivo_id: int):
    """Inhabilita (soft-delete) un motivo de ajuste."""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.MotivosAjuste SET Activo = 0 WHERE MotivoID = ?", (motivo_id,))
        conn.commit()
        return {"message": "Motivo inhabilitado."}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.post("/api/procurement/abonos-batch", response_model=None)
async def registrar_abonos_batch(
    pagos_json: str = Form(...),
    NotificarCorreo: bool = Form(False),
    MontoTotalPagado: float = Form(0.0),
    archivos: List[UploadFile] = File(None)
):
    """Register multiple payments in one transaction and send one consolidated email."""
    import json as _json
    try:
        pagos = _json.loads(pagos_json)
        if not isinstance(pagos, list) or len(pagos) == 0:
            raise HTTPException(status_code=400, detail="Se requiere al menos un pago.")
        
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        filepaths = []
        if archivos:
            os.makedirs("static/uploads", exist_ok=True)
            for archivo in archivos:
                if archivo.filename:
                    ext = os.path.splitext(archivo.filename)[1]
                    filename = f"{uuid.uuid4()}{ext}"
                    filepath = f"static/uploads/{filename}"
                    with open(filepath, "wb") as buffer:
                        shutil.copyfileobj(archivo.file, buffer)
                    filepaths.append(filepath)
        
        rutas_json = _json.dumps(filepaths) if filepaths else ""
        
        # cursor.execute("BEGIN TRANSACTION;") # Rely on pyodbc autocommit=False default
        for p in pagos:
            # Idempotency / Double payment check
            cursor.execute("SELECT Saldo FROM EnterpriseAdmin_AMC.dbo.SAACXP WITH (NOLOCK) WHERE NumeroD = ? AND TipoCxP = '10' AND CodProv = ?", (p['NumeroD'], p['CodProv']))
            saldo_row = cursor.fetchone()
            if not saldo_row or float(saldo_row[0]) <= 0:
                if float(p.get('MontoBsAbonado', 0)) > 0:
                    raise HTTPException(status_code=400, detail=f"La factura {p['NumeroD']} ya no presenta saldo pendiente en el ERP.")
                    
            aplica_idx = 1 if p.get('AplicaIndexacion') in [True, 'true', 'True', 1] else 0
            
            # Phase 8: Lookup mirror fields for independence
            cursor.execute("SELECT Factor, MontoMEx FROM dbo.SACOMP WITH (NOLOCK) WHERE NumeroD = ? AND TipoCom = 'H'", (p['NumeroD'],))
            sacomp_row = cursor.fetchone()
            tasa_orig = float(sacomp_row[0]) if sacomp_row and sacomp_row[0] else None
            mto_orig = float(sacomp_row[1]) if sacomp_row and sacomp_row[1] else None

            cursor.execute("SELECT ISNULL(MAX(AbonoID), 0) FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos WITH (UPDLOCK)")
            max_abono_id = int(cursor.fetchone()[0]) + 1
            
            cursor.execute("""
                INSERT INTO EnterpriseAdmin_AMC.dbo.CxP_Abonos 
                (AbonoID, NumeroD, CodProv, FechaAbono, MontoBsAbonado, TasaCambioDiaAbono, MontoUsdAbonado, AplicaIndexacion, Referencia, RutaComprobante, NotificarCorreo, TasaCambioOrig, MontoMExOrig, TipoAbono, AfectaSaldo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'PAGO_MANUAL', 1)
            """, (
                max_abono_id, p['NumeroD'], p['CodProv'], p['FechaAbono'],
                float(p.get('MontoBsAbonado', 0)), float(p.get('TasaCambioDiaAbono', 0)),
                float(p.get('MontoUsdAbonado', 0)), aplica_idx,
                p.get('Referencia', ''), rutas_json, 1 if NotificarCorreo else 0,
                tasa_orig, mto_orig
            ))
            
            monto_abono = float(p.get('MontoBsAbonado', 0))
            monto_ajuste = float(p.get('MontoAjusteBs', 0))
            
            # Phase 14: Log Discounts in Batch (Feature Parity)
            monto_desc = float(p.get('MontoDescuentoBs', 0))
            if monto_desc > 0:
                cursor.execute("SELECT ISNULL(MAX(AbonoID), 0) FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos")
                max_desc_id = int(cursor.fetchone()[0]) + 1
                cursor.execute("""
                    INSERT INTO EnterpriseAdmin_AMC.dbo.CxP_Abonos 
                    (AbonoID, NumeroD, CodProv, FechaAbono, MontoBsAbonado, TasaCambioDiaAbono, MontoUsdAbonado,
                     AplicaIndexacion, Referencia, RutaComprobante, NotificarCorreo, TasaCambioOrig, MontoMExOrig, TipoAbono, MotivoAjusteID, AfectaSaldo)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'DESCUENTO APLICADO (LOTE)', NULL, 0, ?, ?, 'DESCUENTO', ?, 0)
                """, (
                    max_desc_id, p['NumeroD'], p['CodProv'], p['FechaAbono'],
                    monto_desc, float(p.get('TasaCambioDiaAbono', 0)),
                    monto_desc / float(p.get('TasaCambioDiaAbono', 1)) if float(p.get('TasaCambioDiaAbono', 0)) > 0 else 0,
                    aplica_idx, tasa_orig, mto_orig, p.get('MotivoDescuentoID')
                ))

            monto_desc_base = float(p.get('MontoDescuentoBaseBs', 0))
            if monto_desc_base > 0:
                cursor.execute("SELECT ISNULL(MAX(AbonoID), 0) FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos")
                max_desc_base_id = int(cursor.fetchone()[0]) + 1
                cursor.execute("""
                    INSERT INTO EnterpriseAdmin_AMC.dbo.CxP_Abonos 
                    (AbonoID, NumeroD, CodProv, FechaAbono, MontoBsAbonado, TasaCambioDiaAbono, MontoUsdAbonado,
                     AplicaIndexacion, Referencia, RutaComprobante, NotificarCorreo, TasaCambioOrig, MontoMExOrig, TipoAbono, MotivoAjusteID, AfectaSaldo)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'DESCUENTO BASE (LOTE)', NULL, 0, ?, ?, 'DESCUENTO_BASE', NULL, 0)
                """, (
                    max_desc_base_id, p['NumeroD'], p['CodProv'], p['FechaAbono'],
                    monto_desc_base, float(p.get('TasaCambioDiaAbono', 0)),
                    monto_desc_base / float(p.get('TasaCambioDiaAbono', 1)) if float(p.get('TasaCambioDiaAbono', 0)) > 0 else 0,
                    aplica_idx, tasa_orig, mto_orig
                ))

            if monto_ajuste > 0:
                monto_abono += monto_ajuste
                cursor.execute("SELECT ISNULL(MAX(AbonoID), 0) FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos")
                max_abono_id2 = int(cursor.fetchone()[0]) + 1
                tasa_dia = float(p.get('TasaCambioDiaAbono', 0))
                motivo_ajuste_id = p.get('MotivoAjusteID') or None
                cursor.execute("""
                    INSERT INTO EnterpriseAdmin_AMC.dbo.CxP_Abonos 
                    (AbonoID, NumeroD, CodProv, FechaAbono, MontoBsAbonado, TasaCambioDiaAbono, MontoUsdAbonado, AplicaIndexacion, Referencia, RutaComprobante, NotificarCorreo, TasaCambioOrig, MontoMExOrig, TipoAbono, MotivoAjusteID, AfectaSaldo)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'AJUSTE DE SISTEMA', NULL, 0, ?, ?, 'AJUSTE', ?, 1)
                """, (
                    max_abono_id2, p['NumeroD'], p['CodProv'], p['FechaAbono'],
                    monto_ajuste, tasa_dia, monto_ajuste / tasa_dia if tasa_dia > 0 else 0,
                    aplica_idx, tasa_orig, mto_orig, motivo_ajuste_id
                ))

            
            # Update SACOMP.MtoPagos and SAACXP.Saldo for consistency
            nro_unico = p.get('NroUnico')
            if nro_unico:
                cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SACOMP SET MtoPagos = ISNULL(MtoPagos, 0) + ? WHERE NroUnico = (SELECT TOP 1 c.NroUnico FROM EnterpriseAdmin_AMC.dbo.SACOMP c JOIN EnterpriseAdmin_AMC.dbo.SAACXP x ON c.NumeroD = x.NumeroD AND c.CodProv = x.CodProv WHERE x.NroUnico = ?)", (monto_abono, nro_unico))
                cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAACXP SET Saldo = Saldo - ? WHERE NroUnico = ?", (monto_abono, nro_unico))
            else:
                cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SACOMP SET MtoPagos = ISNULL(MtoPagos, 0) + ? WHERE NumeroD = ? AND CodProv = ?", (monto_abono, p['NumeroD'], p['CodProv']))
                cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAACXP SET Saldo = Saldo - ? WHERE NumeroD = ? AND TipoCxP = '10' AND CodProv = ?", (monto_abono, p['NumeroD'], p['CodProv']))
            
            # Also update SAPROV.Saldo (Provider global balance tracking)
            cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAPROV SET Saldo = Saldo - ? WHERE CodProv = ?", (monto_abono, p['CodProv']))
        
        # Phase 5: Excess payment as Credit Note
        total_abonos = sum(float(p.get('MontoBsAbonado', 0)) for p in pagos)
        excedente = MontoTotalPagado - total_abonos
        if excedente > 0.01:
            # Create Credit Note Request using first invoice as reference
            ref_p = pagos[0]
            tasa = float(ref_p.get('TasaCambioDiaAbono', 0))
            cursor.execute("SELECT Factor, MontoMEx FROM dbo.SACOMP WITH (NOLOCK) WHERE NumeroD = ? AND TipoCom = 'H'", (ref_p['NumeroD'],))
            sacomp_row = cursor.fetchone()
            tasa_orig = float(sacomp_row[0]) if sacomp_row and sacomp_row[0] else None
            mto_orig = float(sacomp_row[1]) if sacomp_row and sacomp_row[1] else None

            cursor.execute("""
                INSERT INTO EnterpriseAdmin_AMC.Procurement.CreditNotesTracking
                (CodProv, NumeroD, Motivo, MontoBs, TasaCambio, MontoUsd, Estatus, Observacion, TasaCambioOrig, MontoMExOrig)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                ref_p['CodProv'], 
                ref_p['NumeroD'],
                'PAGO_EXCESO',
                excedente,
                tasa,
                excedente / tasa if tasa > 0 else 0,
                'PENDIENTE',
                f"Excedente de pago lote ({len(pagos)} facturas)",
                tasa_orig, mto_orig
            ))

        
        # Phase 11: Auto-generate Debit Note for Indexation Discrepancies
        for p in pagos:
            if p.get('AplicaIndexacion') in [True, 'true', 'True', 1]:
                cursor.execute("SELECT Monto FROM EnterpriseAdmin_AMC.dbo.SAACXP WITH (NOLOCK) WHERE NumeroD = ? AND TipoCxP = '10' AND CodProv = ?", (p['NumeroD'], p['CodProv']))
                saacxp_row = cursor.fetchone()
                if saacxp_row:
                    orig_bs = float(saacxp_row[0])
                    cursor.execute("SELECT SUM(MontoBsAbonado) FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos WITH (NOLOCK) WHERE NumeroD = ? AND CodProv = ? AND ISNULL(AfectaSaldo, 1) = 1", (p['NumeroD'], p['CodProv']))
                    total_paid_bs = cursor.fetchone()[0] or 0
                    
                    discrepancia = float(total_paid_bs) - orig_bs
                    if discrepancia > 0.1: # Threshold to avoid cents noise
                        # Check existance to avoid duplicates
                        cursor.execute("SELECT COUNT(*) FROM EnterpriseAdmin_AMC.Procurement.DebitNotesTracking WHERE NumeroD = ? AND CodProv = ? AND Motivo = 'INDEXACION' AND Estatus = 'PENDIENTE'", (p['NumeroD'], p['CodProv']))
                        if cursor.fetchone()[0] == 0:
                            tasa_pago = float(p.get('TasaCambioDiaAbono', 0))
                            cursor.execute("SELECT Factor, MontoMEx FROM dbo.SACOMP WITH (NOLOCK) WHERE NumeroD = ? AND TipoCom = 'H'", (p['NumeroD'],))
                            scomp = cursor.fetchone()
                            t_orig = float(scomp[0]) if scomp and scomp[0] else None
                            m_orig = float(scomp[1]) if scomp and scomp[1] else None
                            
                            cursor.execute("""
                                INSERT INTO EnterpriseAdmin_AMC.Procurement.DebitNotesTracking 
                                (CodProv, NumeroD, Motivo, MontoBs, TasaCambio, MontoUsd, Estatus, Observacion, TasaCambioOrig, MontoMExOrig)
                                VALUES (?, ?, 'INDEXACION', ?, ?, ?, 'PENDIENTE', ?, ?, ?)
                            """, (
                                p['CodProv'], p['NumeroD'], discrepancia, tasa_pago,
                                discrepancia / tasa_pago if tasa_pago > 0 else 0,
                                f"Auto-generado por indexación (Lote)", t_orig, m_orig
                            ))

        conn.commit()
        
        email_sent = False
        if NotificarCorreo:
            cod_prov = pagos[0]['CodProv']
            cursor.execute("""
                SELECT c.Email, p.Descrip 
                FROM EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones c 
                LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON c.CodProv = p.CodProv 
                WHERE c.CodProv = ?
            """, (cod_prov,))
            row = cursor.fetchone()
            if row and row.Email:
                pago_data_list = []
                for p in pagos:
                    pago_data_list.append({
                        "NumeroD": p['NumeroD'],
                        "CodProv": p['CodProv'],
                        "FechaAbono": p.get('FechaAbono', ''),
                        "MontoBsAbonado": float(p.get('MontoBsAbonado', 0)),
                        "MontoUsdAbonado": float(p.get('MontoUsdAbonado', 0)),
                        "TasaCambioDiaAbono": float(p.get('TasaCambioDiaAbono', 0)),
                        "AplicaIndexacion": "Sí" if p.get('AplicaIndexacion') in [True, 'true', 'True', 1] else "No",
                        "Referencia": p.get('Referencia', '')
                    })
                email_sent = safe_send_email(
                    row.Email, row.Descrip or "Proveedor",
                    f"{len(pagos)}_Facturas", pago_data_list, filepath
                )
        
        logging.info(f"Batch abonos: {len(pagos)} pagos registrados")
        return {"message": f"{len(pagos)} pagos registrados exitosamente.", "count": len(pagos), "email_sent": email_sent}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        with open(r"C:\source\CuentasPorPagar\patch_debug.log", "a", encoding="utf-8") as rf:
            rf.write(f"BATCH ERROR: {str(e)}\n{traceback.format_exc()}\n")
        logging.error(f"Error batch abonos: {e}", exc_info=True)
        if 'conn' in locals():
            conn.rollback()
        raise HTTPException(status_code=500, detail="Fallo en la base de datos al procesar el lote de pagos. Transacción revertida completamente.")
    finally:
        if 'conn' in locals():
            conn.close()

@router.post("/api/procurement/abonos")
async def registrar_abono(
    NumeroD: str = Form(...),
    CodProv: str = Form(...),
    FechaAbono: str = Form(...),
    MontoBsAbonado: float = Form(...),
    TasaCambioDiaAbono: float = Form(...),
    MontoUsdAbonado: float = Form(...),
    AplicaIndexacion: str = Form(...),
    Referencia: str = Form(""),
    NotificarCorreo: bool = Form(False),
    MontoTotalPagado: float = Form(0.0),
    force_send: bool = Form(False),
    archivos: List[UploadFile] = File(None),
    NroUnico: Optional[int] = Form(None),
    MontoAjusteBs: Optional[float] = Form(None),
    MotivoAjusteID: Optional[int] = Form(None),
    MontoDescuentoBs: Optional[float] = Form(None),
    MotivoDescuentoID: Optional[int] = Form(None),
    MontoDescuentoBaseBs: Optional[float] = Form(None)
):
    try:
        logging.info(f"REGISTRAR ABONO CALLED - NumeroD: {NumeroD}, MontoBs: {MontoBsAbonado}, MontoAjusteBs: {MontoAjusteBs}, MontoDescuentoBaseBs: {MontoDescuentoBaseBs}")
        conn = database.get_db_connection()
        cursor = conn.cursor()

        aplica_idx = 1 if AplicaIndexacion.lower() == 'true' else 0
        notificar = 1 if NotificarCorreo else 0
        force = 1 if force_send else 0

        import json as _json
        filepaths = []
        if archivos:
            os.makedirs("static/uploads", exist_ok=True)
            for archivo in archivos:
                if archivo.filename:
                    ext = os.path.splitext(archivo.filename)[1]
                    filename = f"{uuid.uuid4()}{ext}"
                    filepath = f"static/uploads/{filename}"
                    with open(filepath, "wb") as buffer:
                        shutil.copyfileobj(archivo.file, buffer)
                    filepaths.append(filepath)

        rutas_json = _json.dumps(filepaths) if filepaths else ""

        # Idempotency / Double payment check
        if NroUnico:
            cursor.execute("SELECT Saldo FROM EnterpriseAdmin_AMC.dbo.SAACXP WITH (NOLOCK) WHERE NroUnico = ?", (NroUnico,))
        else:
            cursor.execute("SELECT Saldo FROM EnterpriseAdmin_AMC.dbo.SAACXP WITH (NOLOCK) WHERE NumeroD = ? AND TipoCxP = '10' AND CodProv = ?", (NumeroD, CodProv))
            
        saldo_row = cursor.fetchone()
        if not saldo_row or float(saldo_row[0]) <= 0:
            if float(MontoBsAbonado) > 0:
                raise HTTPException(status_code=400, detail=f"La factura {NumeroD} ya no presenta saldo pendiente en el ERP.")

        # Phase 8: Mirror fields
        cursor.execute("SELECT Factor, MontoMEx FROM dbo.SACOMP WITH (NOLOCK) WHERE NumeroD = ? AND TipoCom = 'H'", (NumeroD,))
        sacomp_row = cursor.fetchone()
        tasa_orig = float(sacomp_row[0]) if sacomp_row and sacomp_row[0] else None
        mto_orig = float(sacomp_row[1]) if sacomp_row and sacomp_row[1] else None

        cursor.execute("SELECT ISNULL(MAX(AbonoID), 0) FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos WITH (UPDLOCK)")
        max_abono_id = int(cursor.fetchone()[0]) + 1

        insert_query = """
            INSERT INTO EnterpriseAdmin_AMC.dbo.CxP_Abonos 
            (AbonoID, NumeroD, CodProv, FechaAbono, MontoBsAbonado, TasaCambioDiaAbono, MontoUsdAbonado, AplicaIndexacion, Referencia, RutaComprobante, NotificarCorreo, TasaCambioOrig, MontoMExOrig, TipoAbono, AfectaSaldo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'PAGO_MANUAL', 1)
        """
        cursor.execute(insert_query, (
            max_abono_id, NumeroD, CodProv, FechaAbono,
            MontoBsAbonado, TasaCambioDiaAbono,
            MontoUsdAbonado, aplica_idx, Referencia,
            rutas_json, notificar, tasa_orig, mto_orig
        ))
        
        monto_total_reducir = MontoBsAbonado
        
        if MontoDescuentoBs and MontoDescuentoBs > 0:
            # Registro informativo del descuento aplicado - NO afecta saldo (ya fue restado en la matemática del frontend)
            cursor.execute("SELECT ISNULL(MAX(AbonoID), 0) FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos")
            max_desc_id = int(cursor.fetchone()[0]) + 1
            desc_motivo_id = MotivoDescuentoID or None
            cursor.execute("""
                INSERT INTO EnterpriseAdmin_AMC.dbo.CxP_Abonos 
                (AbonoID, NumeroD, CodProv, FechaAbono, MontoBsAbonado, TasaCambioDiaAbono, MontoUsdAbonado,
                 AplicaIndexacion, Referencia, RutaComprobante, NotificarCorreo, TasaCambioOrig, MontoMExOrig, TipoAbono, MotivoAjusteID, AfectaSaldo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'DESCUENTO APLICADO', NULL, 0, ?, ?, 'DESCUENTO', ?, 0)
            """, (
                max_desc_id, NumeroD, CodProv, FechaAbono,
                MontoDescuentoBs, TasaCambioDiaAbono,
                MontoDescuentoBs / TasaCambioDiaAbono if TasaCambioDiaAbono > 0 else 0,
                aplica_idx, tasa_orig, mto_orig, desc_motivo_id
            ))
            
        if MontoDescuentoBaseBs and MontoDescuentoBaseBs > 0:
            cursor.execute("SELECT ISNULL(MAX(AbonoID), 0) FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos")
            max_desc_base_id = int(cursor.fetchone()[0]) + 1
            cursor.execute("""
                INSERT INTO EnterpriseAdmin_AMC.dbo.CxP_Abonos 
                (AbonoID, NumeroD, CodProv, FechaAbono, MontoBsAbonado, TasaCambioDiaAbono, MontoUsdAbonado,
                 AplicaIndexacion, Referencia, RutaComprobante, NotificarCorreo, TasaCambioOrig, MontoMExOrig, TipoAbono, MotivoAjusteID, AfectaSaldo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'DESCUENTO COMERCIAL BASE', NULL, 0, ?, ?, 'DESCUENTO_BASE', NULL, 0)
            """, (
                max_desc_base_id, NumeroD, CodProv, FechaAbono,
                MontoDescuentoBaseBs, TasaCambioDiaAbono,
                MontoDescuentoBaseBs / TasaCambioDiaAbono if TasaCambioDiaAbono > 0 else 0,
                aplica_idx, tasa_orig, mto_orig
            ))
        
        if MontoAjusteBs and MontoAjusteBs > 0:
            monto_total_reducir += MontoAjusteBs
            cursor.execute("SELECT ISNULL(MAX(AbonoID), 0) FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos")
            max_abono_id2 = int(cursor.fetchone()[0]) + 1
            insert_ajuste = """
                INSERT INTO EnterpriseAdmin_AMC.dbo.CxP_Abonos 
                (AbonoID, NumeroD, CodProv, FechaAbono, MontoBsAbonado, TasaCambioDiaAbono, MontoUsdAbonado, AplicaIndexacion, Referencia, RutaComprobante, NotificarCorreo, TasaCambioOrig, MontoMExOrig, TipoAbono, MotivoAjusteID, AfectaSaldo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'AJUSTE DE SISTEMA', NULL, 0, ?, ?, 'AJUSTE', ?, 1)
            """
            cursor.execute(insert_ajuste, (
                max_abono_id2, NumeroD, CodProv, FechaAbono,
                MontoAjusteBs, TasaCambioDiaAbono,
                MontoAjusteBs / TasaCambioDiaAbono if TasaCambioDiaAbono > 0 else 0,
                aplica_idx, tasa_orig, mto_orig, MotivoAjusteID
            ))
        
        # Update SACOMP.MtoPagos and SAACXP.Saldo for consistency
        if NroUnico:
            cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SACOMP SET MtoPagos = ISNULL(MtoPagos, 0) + ? WHERE NroUnico = (SELECT TOP 1 c.NroUnico FROM EnterpriseAdmin_AMC.dbo.SACOMP c JOIN EnterpriseAdmin_AMC.dbo.SAACXP x ON c.NumeroD = x.NumeroD AND c.CodProv = x.CodProv WHERE x.NroUnico = ?)", (monto_total_reducir, NroUnico))
            cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAACXP SET Saldo = Saldo - ? WHERE NroUnico = ?", (monto_total_reducir, NroUnico))
        else:
            cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SACOMP SET MtoPagos = ISNULL(MtoPagos, 0) + ? WHERE NumeroD = ? AND CodProv = ?", (monto_total_reducir, NumeroD, CodProv))
            cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAACXP SET Saldo = Saldo - ? WHERE NumeroD = ? AND TipoCxP = '10' AND CodProv = ?", (monto_total_reducir, NumeroD, CodProv))
        cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAPROV SET Saldo = Saldo - ? WHERE CodProv = ?", (monto_total_reducir, CodProv))

        # Phase 11: Auto-generate Debit Note for Indexation Discrepancies
        if aplica_idx == 1:
            cursor.execute("SELECT Monto FROM EnterpriseAdmin_AMC.dbo.SAACXP WITH (NOLOCK) WHERE NumeroD = ? AND TipoCxP = '10' AND CodProv = ?", (NumeroD, CodProv))
            saacxp_row = cursor.fetchone()
            if saacxp_row:
                orig_bs = float(saacxp_row[0])
                cursor.execute("SELECT SUM(MontoBsAbonado) FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos WITH (NOLOCK) WHERE NumeroD = ? AND CodProv = ? AND ISNULL(AfectaSaldo, 1) = 1", (NumeroD, CodProv))
                total_paid_bs = cursor.fetchone()[0] or 0
                
                discrepancia = float(total_paid_bs) - orig_bs
                if discrepancia > 0.1:
                    cursor.execute("SELECT COUNT(*) FROM EnterpriseAdmin_AMC.Procurement.DebitNotesTracking WHERE NumeroD = ? AND CodProv = ? AND Motivo = 'INDEXACION' AND Estatus = 'PENDIENTE'", (NumeroD, CodProv))
                    if cursor.fetchone()[0] == 0:
                        cursor.execute("""
                            INSERT INTO EnterpriseAdmin_AMC.Procurement.DebitNotesTracking 
                            (CodProv, NumeroD, Motivo, MontoBs, TasaCambio, MontoUsd, Estatus, Observacion, TasaCambioOrig, MontoMExOrig)
                            VALUES (?, ?, 'INDEXACION', ?, ?, ?, 'PENDIENTE', ?, ?, ?)
                        """, (
                            CodProv, NumeroD, discrepancia, TasaCambioDiaAbono,
                            discrepancia / TasaCambioDiaAbono if TasaCambioDiaAbono > 0 else 0,
                            f"Auto-generado por indexación ({FechaAbono})", tasa_orig, mto_orig
                        ))

        conn.commit()

        # Phase 5: Excess payment as Credit Note
        excedente = MontoTotalPagado - MontoBsAbonado
        if excedente > 0.01:
            try:
                conn_nc = database.get_db_connection()
                cursor_nc = conn_nc.cursor()
                cursor_nc.execute("""
                    INSERT INTO EnterpriseAdmin_AMC.Procurement.CreditNotesTracking
                    (CodProv, NumeroD, Motivo, MontoBs, TasaCambio, MontoUsd, Estatus, Observacion)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    CodProv, NumeroD, 'PAGO_EXCESO',
                    excedente, TasaCambioDiaAbono,
                    excedente / TasaCambioDiaAbono if TasaCambioDiaAbono > 0 else 0,
                    'PENDIENTE',
                    f"Excedente de pago factura {NumeroD}"
                ))
                conn_nc.commit()
                conn_nc.close()
            except Exception as e_nc:
                logging.error(f"Error creating auto-credit note: {e_nc}")

        # Enviar correo si corresponde (notificar o force_send) - DESPUES DEL COMMIT para evitar bloqueos
        if notificar == 1 or force == 1:
            cursor.execute("""
                SELECT c.Email, p.Descrip 
                FROM EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones c 
                LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON c.CodProv = p.CodProv 
                WHERE c.CodProv = ?
            """, (CodProv,))
            row = cursor.fetchone()
            if row and row.Email:
                pago_data = {
                    "NumeroD": NumeroD,
                    "CodProv": CodProv,
                    "FechaAbono": FechaAbono,
                    "MontoBsAbonado": MontoBsAbonado,
                    "MontoUsdAbonado": MontoUsdAbonado,
                    "TasaCambioDiaAbono": TasaCambioDiaAbono,
                    "AplicaIndexacion": "Sí" if AplicaIndexacion.lower() == 'true' else "No",
                    "Referencia": Referencia
                }
                email_sent = enviar_correo_pago(row.Email, row.Descrip or "Proveedor", NumeroD, pago_data, filepath)
                logging.info(f"Email sent flag: {email_sent}")
            else:
                logging.warning(f"No notification sent: missing email or provider logic cod={CodProv}")

        return {"message": "Abono registrado exitosamente.", "email_sent": (notificar == 1 or force == 1)}
    except Exception as e:
        logging.error(f"Error registering abono: {e}", exc_info=True)
        if 'conn' in locals():
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()


@router.post("/api/procurement/invoice-settings")
async def update_invoice_settings(
    NumeroD: str = Form(...),
    CodProv: str = Form(...),
    AplicaIndexacion: bool = Form(...)
):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            IF EXISTS (SELECT 1 FROM EnterpriseAdmin_AMC.Procurement.InvoiceSettings WHERE NumeroD = ? AND CodProv = ?)
            BEGIN
                UPDATE EnterpriseAdmin_AMC.Procurement.InvoiceSettings 
                SET AplicaIndexacion = ?, UpdatedAt = GETDATE()
                WHERE NumeroD = ? AND CodProv = ?
            END
            ELSE
            BEGIN
                INSERT INTO EnterpriseAdmin_AMC.Procurement.InvoiceSettings (NumeroD, CodProv, AplicaIndexacion)
                VALUES (?, ?, ?)
            END
        """, (NumeroD, CodProv, 1 if AplicaIndexacion else 0, NumeroD, CodProv, NumeroD, CodProv, 1 if AplicaIndexacion else 0))
        
        conn.commit()
        return {"data": "ok"}
    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
        logging.error(f"Error updating invoice indexation settings: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@router.delete("/api/procurement/abonos/{id_abono}")
async def eliminar_abono(id_abono: int):
    """Elimina permanentemente un abono manual de la base de datos."""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Verify it exists and get amount to revert balances
        cursor.execute("SELECT TipoAbono, MontoBsAbonado, NumeroD, CodProv FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos WHERE AbonoID = ?", (id_abono,))
        row = cursor.fetchone()
        
        if not row:
            raise HTTPException(status_code=404, detail="Abono no encontrado.")
            
        tipo_abono, monto_bs, num_d, cod_prov = row
        monto_bs = float(monto_bs) if monto_bs else 0.0

        # Restrict deletion to manual payments or undefined (legacy). System generated abonos (retenciones) must be cancelled via their modules.
        if tipo_abono in ['RETENCION_IVA', 'RETENCION_ISLR', 'NOTA_CREDITO']:
            raise HTTPException(status_code=400, detail=f"No se puede eliminar directamente un abono de tipo {tipo_abono}. Anule el documento origen en su módulo interactivo correspondiente.")

        # Revert balances in Saint, ensuring MtoPagos never drops below 0 to avoid balance poisoning from improperly synced external records
        cursor.execute("""
            UPDATE EnterpriseAdmin_AMC.dbo.SACOMP WITH (ROWLOCK) 
            SET 
                MtoPagos = CASE WHEN ISNULL(MtoPagos,0) - ? < 0 THEN 0 ELSE ISNULL(MtoPagos,0) - ? END, 
                SaldoAct = SaldoAct + ? 
            WHERE NumeroD = ? AND CodProv = ?
        """, (monto_bs, monto_bs, monto_bs, num_d, cod_prov))
        cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAACXP WITH (ROWLOCK) SET Saldo = Saldo + ? WHERE NumeroD = ? AND CodProv = ? AND TipoCxP = '10'", (monto_bs, num_d, cod_prov))
        cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAPROV WITH (ROWLOCK) SET Saldo = Saldo + ? WHERE CodProv = ?", (monto_bs, cod_prov))

        cursor.execute("DELETE FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos WHERE AbonoID = ?", (id_abono,))
        conn.commit()
        return {"message": "Abono eliminado y saldo de factura revertido exitosamente."}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error eliminating abono {id_abono}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

# Helper: wrap email send so a network error doesn't crash the payment transaction
def safe_send_email(destinatario: str, proveedor_nombre: str, nro_factura: str, pago_data: dict, filepaths=None) -> bool:
    try:
        return enviar_correo_pago(destinatario, proveedor_nombre, nro_factura, pago_data, filepaths)
    except Exception as e:
        logging.warning(f"Email send failed (likely offline): {e}")
        return False

# Endpoint dedicated only to sending email (no payment insertion)
@router.post("/api/procurement/send-email", response_model=None)
async def send_email_only(
    NumeroD: str = Form(...),
    CodProv: str = Form(...),
    archivos: List[UploadFile] = File(None)
):
    """Send a payment notification email for an existing invoice WITHOUT inserting a new payment record."""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT c.Email, p.Descrip 
            FROM EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones c 
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON c.CodProv = p.CodProv 
            WHERE c.CodProv = ?
        """, (CodProv,))
        row = cursor.fetchone()

        if not row or not row.Email:
            return {"email_sent": False, "message": "Proveedor sin email configurado."}

        # Save attachment if provided
        import json as _json
        filepaths = []
        if archivos:
            os.makedirs("static/uploads", exist_ok=True)
            for archivo in archivos:
                if archivo.filename:
                    ext = os.path.splitext(archivo.filename)[1]
                    filename = f"{uuid.uuid4()}{ext}"
                    filepath = f"static/uploads/{filename}"
                    with open(filepath, "wb") as buffer:
                        shutil.copyfileobj(archivo.file, buffer)
                    filepaths.append(filepath)

        # Fetch last abono for context
        cursor.execute("""
            SELECT TOP 1 FechaAbono, MontoBsAbonado, MontoUsdAbonado, TasaCambioDiaAbono, AplicaIndexacion, Referencia, RutaComprobante
            FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos
            WHERE NumeroD = ? AND CodProv = ?
            ORDER BY FechaAbono DESC
        """, (NumeroD, CodProv))
        abono = cursor.fetchone()

        if abono:
            pago_data = {
                "NumeroD": NumeroD,
                "CodProv": CodProv,
                "FechaAbono": str(abono.FechaAbono).split(" ")[0] if abono.FechaAbono else "-", # type: ignore
                "MontoBsAbonado": abono.MontoBsAbonado or 0,
                "MontoUsdAbonado": abono.MontoUsdAbonado or 0,
                "TasaCambioDiaAbono": abono.TasaCambioDiaAbono or 0,
                "AplicaIndexacion": "Sí" if abono.AplicaIndexacion else "No",
                "Referencia": abono.Referencia or "Re-envio de soporte"
            }
            # Use the original comprobante if no new file was uploaded
            attach = filepaths if filepaths else (str(abono.RutaComprobante) if abono.RutaComprobante else None)
        else:
            pago_data = {
                "NumeroD": NumeroD, "CodProv": CodProv,
                "FechaAbono": "-", "MontoBsAbonado": 0,
                "MontoUsdAbonado": 0, "TasaCambioDiaAbono": 0,
                "AplicaIndexacion": "No", "Referencia": "Re-envio de soporte"
            }
            attach = filepaths

        sent = safe_send_email(row.Email, row.Descrip or "Proveedor", NumeroD, pago_data, attach) # type: ignore
        return {"email_sent": sent, "message": "Correo enviado." if sent else "No se pudo enviar el correo."}
    except Exception as e:
        logging.error(f"Error send-email {NumeroD}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

# PATCH endpoint to edit invoice fields - syncs all Saint tables (SAACXP + SACOMP + SAPROV)
@router.patch("/api/cuentas-por-pagar/{numeroD}", response_model=None)
async def editar_factura(numeroD: str, cod_prov: str = Query(default=''), payload: dict = Body(...)):
    """
    Update invoice fields and keep Saint tables fully aligned.
    Fields: FechaE, FechaI, FechaV, SaldoAct (maps to SAACXP.Saldo + SAPROV.Saldo delta), MontoFacturaBS, MontoFacturaUSD, Notas10.
    cod_prov query param is required to correctly identify the SACOMP row.
    """
    print(f"\n[PATCH] numeroD={numeroD} cod_prov={cod_prov} payload={payload}")
    # Write to a persistent debug file with absolute path
    log_file = r"C:\source\CuentasPorPagar\patch_debug.log"
    try:
        with open(log_file, "a", encoding="utf-8") as f:
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{timestamp}] numeroD={numeroD} cod_prov={cod_prov} payload={payload}\n")
    except Exception as e:
        print(f"Error writing to log: {e}")

    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()

        # Update cod_prov from payload if provided
        if "CodProv" in payload:
            cod_prov = payload["CodProv"]

        # ── 1. SAACXP.Saldo + SAPROV.Saldo (sync using delta) ────────────────
        if "SaldoAct" in payload and payload["SaldoAct"] is not None:
            nuevo_saldo = float(payload["SaldoAct"])
            
            # Fetch previous Saldo to calculate delta
            cursor.execute(
                "SELECT Saldo, CodProv FROM EnterpriseAdmin_AMC.dbo.SAACXP WHERE NumeroD = ? AND TipoCxP = '10'", 
                (numeroD,)
            )
            row = cursor.fetchone()
            
            if row is not None:
                viejo_saldo = float(row.Saldo) if row.Saldo is not None else 0.0
                delta = nuevo_saldo - viejo_saldo
                # Only use cod_prov from DB if we don't have it from payload
                if not cod_prov:
                    cod_prov = row.CodProv
                
                # Update SAACXP.Saldo
                cursor.execute(
                    """UPDATE EnterpriseAdmin_AMC.dbo.SAACXP
                       SET Saldo = ?
                       WHERE NumeroD = ? AND TipoCxP = '10'""",
                    (nuevo_saldo, numeroD)
                )
                
                # Update SAPROV.Saldo by applying the delta
                prov_for_delta = cod_prov or row.CodProv
                if delta != 0 and prov_for_delta:
                    cursor.execute(
                        """UPDATE EnterpriseAdmin_AMC.dbo.SAPROV
                           SET Saldo = Saldo + ?
                           WHERE CodProv = ?""",
                        (delta, prov_for_delta)
                    )

        # ── 2. SAACXP: Monto si cambia MontoFacturaBS (Recalculate Saldo Delta) ───────
        if "MontoFacturaBS" in payload and payload["MontoFacturaBS"] is not None:
            new_monto = float(payload["MontoFacturaBS"])
            query_cond = "NumeroD = ? AND TipoCxP = '10'"
            params_cond = [numeroD]
            if cod_prov:
                query_cond += " AND CodProv = ?"
                params_cond.append(cod_prov)
                
            cursor.execute(f"SELECT Monto, Saldo FROM EnterpriseAdmin_AMC.dbo.SAACXP WHERE {query_cond}", tuple(params_cond))
            current = cursor.fetchone()
            
            if current:
                old_monto = float(current.Monto) if current.Monto is not None else 0.0
                old_saldo = float(current.Saldo) if current.Saldo is not None else 0.0
                monto_delta = new_monto - old_monto
                new_saldo = old_saldo + monto_delta
                
                cursor.execute(
                    f"""UPDATE EnterpriseAdmin_AMC.dbo.SAACXP
                       SET Monto = ?, Saldo = ?
                       WHERE {query_cond}""",
                    tuple([new_monto, new_saldo] + params_cond)
                )
                
                if monto_delta != 0 and cod_prov:
                    cursor.execute(
                        """UPDATE EnterpriseAdmin_AMC.dbo.SAPROV
                           SET Saldo = Saldo + ?
                           WHERE CodProv = ?""",
                        (monto_delta, cod_prov)
                    )
            else:
                cursor.execute(
                    f"""UPDATE EnterpriseAdmin_AMC.dbo.SAACXP
                       SET Monto = ?
                       WHERE {query_cond}""",
                    tuple([new_monto] + params_cond)
                )

        # ── 3. SACOMP: fechas + montos ────────────────────────────────────────
        comp_fields = []
        comp_params = []
        if "FechaE" in payload and payload["FechaE"]:
            comp_fields.append("FechaE = ?")
            comp_params.append(payload["FechaE"])
            cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAACXP SET FechaE = ? WHERE NumeroD = ? AND TipoCxP = '10'", (payload["FechaE"], numeroD))
        if "FechaI" in payload and payload["FechaI"]:
            comp_fields.append("FechaI = ?")
            comp_params.append(payload["FechaI"])
            cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAACXP SET FechaI = ? WHERE NumeroD = ? AND TipoCxP = '10'", (payload["FechaI"], numeroD))
        if "FechaV" in payload and payload["FechaV"]:
            comp_fields.append("FechaV = ?")
            comp_params.append(payload["FechaV"])
            cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAACXP SET FechaV = ? WHERE NumeroD = ? AND TipoCxP = '10'", (payload["FechaV"], numeroD))
            
        # Process new explicit fields (Phase 8 logic)
        for field in ["Factor", "MontoMEx", "TotalPrd", "Descto1", "Descto2", "Fletes", "Contado", "Credito", "MtoTotal"]:
            if field in payload and payload[field] is not None:
                comp_fields.append(f"{field} = ?")
                comp_params.append(float(payload[field]))
                if field == "Factor":
                    # Also update SAACXP.Factor since frontend depends on it
                    cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAACXP SET Factor = ? WHERE NumeroD = ? AND TipoCxP = '10'", (float(payload[field]), numeroD))

        # Backward compatibility for MontoFacturaBS mapping if MtoTotal wasn't explicitly sent
        if "MontoFacturaBS" in payload and payload["MontoFacturaBS"] is not None and "MtoTotal" not in payload:
            monto_bs = float(payload["MontoFacturaBS"])
            comp_fields.append("MtoTotal = ?")
            comp_params.append(monto_bs)
            if "Credito" not in payload:
                comp_fields.append("Credito = ?")
                comp_params.append(monto_bs)

        if "Notas10" in payload:
            notas10_val = str(payload["Notas10"]).strip() if payload["Notas10"] is not None else ""
            if notas10_val == "1":
                comp_fields.append("Notas10 = ?")
                comp_params.append("1")
            elif notas10_val == "0" or notas10_val == "":
                # '0' or 'Sin cambio' means clear the field
                comp_fields.append("Notas10 = NULL")

        if "TGravable" in payload and payload["TGravable"] is not None:
            comp_fields.append("TGravable = ?")
            comp_params.append(float(payload["TGravable"]))

        if "IVA" in payload and payload["IVA"] is not None:
            iva_val = float(payload["IVA"])
            comp_fields.append("MtoTax = ?")
            comp_params.append(iva_val)
            # Update SAACXP.MtoTax too
            cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAACXP SET MtoTax = ? WHERE NumeroD = ? AND TipoCxP = '10'", (iva_val, numeroD))
        elif "MtoTax" in payload and payload["MtoTax"] is not None:
            # Also handle if sent as MtoTax
            iva_val = float(payload["MtoTax"])
            comp_fields.append("MtoTax = ?")
            comp_params.append(iva_val)
            cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAACXP SET MtoTax = ? WHERE NumeroD = ? AND TipoCxP = '10'", (iva_val, numeroD))

        if comp_fields:
            set_clause = ", ".join(comp_fields)
            # Filter by both CodProv and NumeroD to match the JOIN used in SELECT
            logging.info(f"[PATCH SACOMP] numeroD={numeroD} cod_prov={cod_prov!r} fields={comp_fields} params={comp_params}")
            if cod_prov:
                comp_params.append(cod_prov)
                comp_params.append(numeroD)
                logging.info(f"[PATCH SACOMP] SQL: UPDATE SACOMP SET {set_clause} WHERE CodProv=? AND NumeroD=?  params={tuple(comp_params)}")
                cursor.execute(
                    f"UPDATE EnterpriseAdmin_AMC.dbo.SACOMP SET {set_clause} WHERE CodProv = ? AND NumeroD = ?",
                    tuple(comp_params)
                )
                logging.info(f"[PATCH SACOMP] rowcount={cursor.rowcount}")
            else:
                comp_params.append(numeroD) # type: ignore
                cursor.execute(
                    f"UPDATE EnterpriseAdmin_AMC.dbo.SACOMP SET {set_clause} WHERE NumeroD = ?",
                    tuple(comp_params)
                )
            
            print(f"[PATCH SACOMP] Affected Rows: {cursor.rowcount}")
            with open(log_file, "a", encoding="utf-8") as f:
                f.write(f"[PATCH SACOMP] Affected Rows: {cursor.rowcount}\n")

        conn.commit()
        return {"message": f"Factura {numeroD} actualizada y tablas Saint sincronizadas correctamente."}
    except Exception as e:
        logging.error(f"Error updating factura {numeroD}: {e}", exc_info=True)
        if 'conn' in locals():
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@router.get("/api/procurement/cxp-status")
async def get_cxp_status(cod_prov: str = Query(...), numero_d: str = Query(...), nro_unico: Optional[int] = Query(None)):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # We need to fetch:
        # 1. Invoice details from SACOMP/SAACXP (Monto, Saldo, Fechas, Tasa original)
        # 2. Provider conditions from ProveedorCondiciones
        # 3. Sum of all Abonos from CxP_Abonos
        
        where_clause = "cxp.CodProv = ? AND cxp.NumeroD = ? AND cxp.TipoCxP = '10'"
        params = [cod_prov, numero_d]
        if nro_unico:
            where_clause = "cxp.NroUnico = ?"
            params = [nro_unico]

        query = f"""
            SELECT 
                cxp.NumeroD, cxp.CodProv, cxp.Monto, cxp.Saldo, 
                cxp.FechaE, cxp.FechaV AS FechaVSaint,
                comp.FechaI, comp.Notas10,
                comp.TGravable, cxp.MtoTax,
                comp.Factor, comp.MontoMEx,
                ISNULL(cond.DiasNoIndexacion, 0) AS DiasNoIndexacion,
                ISNULL(cond.IndexaIVA, 1) AS IndexaIVA,
                ISNULL(cond.BaseDiasCredito, 'EMISION') AS BaseDiasCredito,
                ISNULL(cond.DiasVencimiento, prov.diascred) AS DiasVencimiento,
                ISNULL(cond.DecimalesTasa, 4) AS DecimalesTasa,
                ISNULL(cond.DescuentoBase_Pct, 0) AS DescuentoBase_Pct,
                ISNULL(cond.DescuentoBase_Condicion, 'INDEPENDIENTE') AS DescuentoBase_Condicion,
                ISNULL(cond.DescuentoBase_DeduceIVA, 0) AS DescuentoBase_DeduceIVA,
                prov.Descrip AS ProveedorNombre,
                prov.NumeroUP, prov.FechaUP, prov.MontoUP,
                dt_emision.dolarbcv AS TasaEmisionCalculada,
                CASE WHEN ISNULL(cxp.Factor, 0) > 1 THEN cxp.Factor ELSE dt_emision.dolarbcv END AS TasaEmision,
                ISNULL(abonos.TotalUsdAbonado, 0) AS TotalUsdAbonado,
                ISNULL(abonos.TotalBsAbonado, 0) AS TotalBsAbonado,
                ISNULL(abonos.TotalIVA, 0) AS RetencionIvaAbonada,
                ISNULL(abonos.TotalISLR, 0) AS RetencionIslrAbonada,
                ISNULL(prov.PorctRet, 0) AS PorctRet,
                ISNULL(prov.EsReten, 0) AS EsReten,
                prov.ID3 AS RIF,
                -- Effective TipoPersona: LOCAL override wins, then SAINT (ID3 prefix), then NULL
                COALESCE(
                    cond.TipoPersona,
                    CASE WHEN LEFT(LTRIM(ISNULL(prov.ID3,'')),1) IN ('J','G','C') THEN 'PJ'
                         WHEN LEFT(LTRIM(ISNULL(prov.ID3,'')),1) IN ('V','E','P') THEN 'PN'
                         ELSE NULL END
                ) AS TipoPersona,
                cond.TipoPersona AS TipoPersonaLocal,
                inv_set.AplicaIndexacion AS AplicaIndexacionOverride
            FROM dbo.SAACXP cxp
            LEFT JOIN dbo.SACOMP comp ON cxp.CodProv = comp.CodProv AND cxp.NumeroD = comp.NumeroD
            LEFT JOIN dbo.SAPROV prov ON cxp.CodProv = prov.CodProv
            LEFT JOIN EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones cond ON cxp.CodProv = cond.CodProv
            LEFT JOIN EnterpriseAdmin_AMC.Procurement.InvoiceSettings inv_set ON cxp.CodProv = inv_set.CodProv AND cxp.NumeroD = inv_set.NumeroD
            OUTER APPLY (
                SELECT SUM(MontoUsdAbonado) as TotalUsdAbonado, SUM(MontoBsAbonado) as TotalBsAbonado,
                       SUM(CASE WHEN TipoAbono = 'RETENCION_IVA' THEN MontoBsAbonado ELSE 0 END) AS TotalIVA,
                       SUM(CASE WHEN TipoAbono = 'RETENCION_ISLR' THEN MontoBsAbonado ELSE 0 END) AS TotalISLR
                FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos
                WHERE CodProv = cxp.CodProv AND NumeroD = cxp.NumeroD AND ISNULL(AfectaSaldo, 1) = 1
            ) abonos
            OUTER APPLY (
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday 
                WHERE CAST(fecha AS DATE) <= CAST(cxp.FechaE AS DATE)
                ORDER BY fecha DESC
            ) dt_emision
            OUTER APPLY (
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday 
                WHERE dolarbcv IS NOT NULL
                ORDER BY id DESC
            ) dt_actual
            WHERE {where_clause}
        """
        cursor.execute(query, params)
        row = cursor.fetchone()
        
        if not row:
            raise HTTPException(status_code=404, detail="Factura no encontrada")
            
        columns = [column[0] for column in cursor.description]
        data = dict(zip(columns, row))

        # Fetch Settings
        cursor.execute("SELECT SettingKey, SettingValue FROM EnterpriseAdmin_AMC.Procurement.Settings")
        settings = {srow[0]: srow[1] for srow in cursor.fetchall()}
        tasa_source = settings.get('TasaEmisionSource', 'DOLARTODAY')
        usd_source = settings.get('MontoUSDSource', 'CALCULATED')
        islr_persona_source = settings.get('ISLRPersonaSource', 'SAINT')
        data['ISLRPersonaSource'] = islr_persona_source

        # Override TasaEmision if configured to parse from Saint
        if tasa_source == 'SACOMP_FACTOR' and data.get('Factor') and data['Factor'] > 0:
            data['TasaEmision'] = float(data['Factor'])
        
        # Calculate dynamic dates
        from datetime import datetime, timedelta
        
        # Base date for calculations
        base_date_str = data['FechaE'] if data['BaseDiasCredito'] == 'EMISION' else (data['FechaI'] or data['FechaE'])
        # base_date_str comes as string or datetime from DB (pyodbc parses to datetime)
        if isinstance(base_date_str, str):
            base_date = datetime.strptime(base_date_str[:10], '%Y-%m-%d')
        else:
            base_date = base_date_str
            
        if data.get('Notas10') == '1':
            fecha_ni = base_date
        else:
            fecha_ni = base_date + timedelta(days=int(data['DiasNoIndexacion']))
            
        fecha_v = base_date + timedelta(days=int(data['DiasVencimiento']))
        
        data['FechaNI_Calculada'] = fecha_ni.strftime('%Y-%m-%d')
        data['FechaV_Calculada'] = fecha_v.strftime('%Y-%m-%d')
        
        # Convert numeric types to float for JSON
        for k, v in data.items():
            if hasattr(v, 'quantize') or isinstance(v, float): # Decimal or float
                data[k] = float(v) if v is not None else 0.0
            elif isinstance(v, datetime): # type: ignore
                data[k] = v.strftime('%Y-%m-%d')
                
        # Compute specific fields based on settings
        monto_original_usd = 0
        if usd_source == 'SACOMP_MONOMEX' and data.get('MontoMEx') and data['MontoMEx'] > 0:
            monto_original_usd = float(data['MontoMEx'])
        else:
            monto_original_usd = data['Monto'] / data['TasaEmision'] if data['TasaEmision'] else 0
            
        saldo_usd = monto_original_usd - data['TotalUsdAbonado']
        data['MontoOriginalUSD'] = round(monto_original_usd, 4)
        data['SaldoRestanteUSD'] = round(saldo_usd, 4)
        # Note: DiferenciaBs is typically calculated on the frontend before closing an invoice 
        # based on exactly what was paid vs what the invoice was worth originally,
        # but the backend provides all necessary totals.

        # Fetch Abonos History
        history_query = """
            SELECT
                ab.AbonoID, ab.FechaAbono, ab.MontoBsAbonado, ab.TasaCambioDiaAbono,
                ab.MontoUsdAbonado, ab.AplicaIndexacion, ab.Referencia, ab.TipoAbono,
                ab.AfectaSaldo,
                ISNULL(ma.Descripcion, ab.Referencia) AS DescripcionAjuste,
                ma.Codigo AS CodigoMotivo
            FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos ab
            LEFT JOIN EnterpriseAdmin_AMC.Procurement.MotivosAjuste ma
                ON ab.MotivoAjusteID = ma.MotivoID
            WHERE ab.CodProv = ? AND ab.NumeroD = ?
            ORDER BY ab.FechaAbono ASC, ab.AbonoID ASC
        """
        cursor.execute(history_query, (cod_prov, numero_d))
        history_cols = [column[0] for column in cursor.description]
        history_data = [dict(zip(history_cols, row)) for row in cursor.fetchall()]
        
        for record in history_data:
            for k, v in record.items():
                if hasattr(v, 'quantize') or isinstance(v, float):
                    record[k] = float(v)
                elif isinstance(v, datetime): # type: ignore
                    record[k] = v.strftime('%Y-%m-%d')
                    
        data['HistorialAbonos'] = history_data

        # Fetch dynamic discounts
        cursor.execute("SELECT DiasDesde, DiasHasta, Porcentaje, DeduceIVA FROM EnterpriseAdmin_AMC.Procurement.ProveedorDescuentosProntoPago WITH (NOLOCK) WHERE CodProv = ? ORDER BY DiasDesde ASC", (cod_prov,))
        desc_cols = [c[0] for c in cursor.description]
        data['Descuentos'] = [dict(zip(desc_cols, row)) for row in cursor.fetchall()]

        return {"data": data}
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching cxp status: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@router.get("/api/procurement/debit-notes")
async def get_debit_notes(search: Optional[str] = None, estatus: Optional[str] = None):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # We need invoices where Total amount paid (Bs) > original invoice amount (Bs)
        # We join Cxp_Abonos and SAACXP.
        query = """
            SELECT 
                cxp.CodProv, 
                prov.Descrip AS ProveedorNombre,
                cxp.NumeroD, 
                cxp.FechaE AS FechaEmision,
                cxp.Monto AS MontoOriginalBs,
                ISNULL(abonos.TotalBsAbonado, 0) AS TotalBsAbonado,
                ISNULL(abonos.TotalBsAbonado, 0) - cxp.Monto AS MontoNotaDebitoBs,
                ISNULL(dnt.Estatus, 'PENDIENTE') AS Estatus,
                dnt.NotaDebitoID,
                dnt.MontoRetencionBs AS DB_MontoRetencionBs,
                COALESCE(cond.Email, prov.Email, '') AS Email,
                ISNULL(prov.PorctRet, 0) AS PorctRet
            FROM dbo.SAACXP cxp
            INNER JOIN dbo.SAPROV prov ON cxp.CodProv = prov.CodProv
            LEFT JOIN EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones cond ON cxp.CodProv = cond.CodProv
            LEFT JOIN EnterpriseAdmin_AMC.Procurement.DebitNotesTracking dnt ON cxp.CodProv = dnt.CodProv AND cxp.NumeroD = dnt.NumeroD
            CROSS APPLY (
                SELECT SUM(a.MontoBsAbonado) as TotalBsAbonado
                FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos a
                WHERE a.CodProv = cxp.CodProv AND a.NumeroD = cxp.NumeroD AND ISNULL(a.AfectaSaldo, 1) = 1
            ) abonos
        """
        params = []
        where_clauses = [
            "cxp.TipoCxP = '10'",
            "ISNULL(abonos.TotalBsAbonado, 0) > cxp.Monto + 0.1",
            # Excluir facturas donde se haya registrado un ajuste con AnulaNotaDebito=1
            "NOT EXISTS (SELECT 1 FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos ab2"
            " INNER JOIN EnterpriseAdmin_AMC.Procurement.MotivosAjuste ma ON ab2.MotivoAjusteID = ma.MotivoID"
            " WHERE ab2.NumeroD = cxp.NumeroD AND ab2.CodProv = cxp.CodProv AND ma.AnulaNotaDebito = 1)"
        ]

        if search:
            where_clauses.append("(cxp.CodProv LIKE ? OR prov.Descrip LIKE ? OR cxp.NumeroD LIKE ?)")
            search_val = f"%{search}%"
            params.extend([search_val, search_val, search_val])
        
        if estatus:
            where_clauses.append("ISNULL(dnt.Estatus, 'PENDIENTE') = ?")
            params.append(estatus)

        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)

        query += " ORDER BY prov.Descrip ASC, cxp.FechaE DESC"
        
        cursor.execute(query, tuple(params))
        columns = [column[0] for column in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Convert Decimals
        for r in results:
            r['MontoOriginalBs'] = float(r['MontoOriginalBs'] or 0)
            r['TotalBsAbonado'] = float(r['TotalBsAbonado'] or 0)
            r['MontoNotaDebitoBs'] = float(r['MontoNotaDebitoBs'] or 0)
            r['PorctRet'] = float(r['PorctRet'] or 0)
            
            if r['Estatus'] == 'EMITIDA' and r['DB_MontoRetencionBs'] is not None:
                r['MontoRetencionBs'] = float(r['DB_MontoRetencionBs'])
            else:
                # Estimar Retención (usando 16% base IVA)
                r['MontoRetencionBs'] = float(r['MontoNotaDebitoBs'] * 0.16 * (r['PorctRet'] / 100.0))
            
            r.pop('DB_MontoRetencionBs', None)
            
        return {"data": results}
    except Exception as e:
        logging.error(f"Error fetching debit notes: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@router.post("/api/procurement/debit-notes/send-request")
async def send_debit_note_requests(payload: DebitNoteActionRequest):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # In a real-world scenario, here we would:
        # 1. Group payload.Invoices by CodProv.
        # 2. Iterate each provider, fetch their Email.
        # 3. Generate an Excel attachment /api/export/debit-notes?cod_prov=...
        # 4. Use python standard smtplib using os.getenv("SMTP_USERNAME") and os.getenv("SMTP_PASSWORD")
        # 5. Send email. 
        # For now, we will mark them as SOLICITUD_ENVIADA to complete the process tracking for the user.
        
        for inv in payload.Invoices:
            # Check if exists in tracking table
            cursor.execute("SELECT CodProv FROM EnterpriseAdmin_AMC.Procurement.DebitNotesTracking WHERE CodProv=? AND NumeroD=?", (inv.CodProv, inv.NumeroD))
            if cursor.fetchone():
                cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.DebitNotesTracking SET Estatus='SOLICITUD_ENVIADA', FechaSolicitud=GETDATE() WHERE CodProv=? AND NumeroD=?", (inv.CodProv, inv.NumeroD))
            else:
                cursor.execute("INSERT INTO EnterpriseAdmin_AMC.Procurement.DebitNotesTracking (CodProv, NumeroD, Estatus, FechaSolicitud) VALUES (?, ?, 'SOLICITUD_ENVIADA', GETDATE())", (inv.CodProv, inv.NumeroD))
        
        conn.commit()
        return {"message": f"Se han marcado {len(payload.Invoices)} facturas como Solicitud Enviada."}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error send_debit_note_requests: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.post("/api/procurement/debit-notes/register")
async def register_debit_note(payload: DebitNoteRegisterRequest):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        nd_num = payload.NotaDebitoID
        nd_ctrl = payload.ControlID or nd_num
        
        for inv in payload.Invoices:
            codprov = inv.CodProv
            fac_num = inv.NumeroD
            
            # 1. Fetch exact surplus (Negative Balance) from Invoice in SACOMP/SAACXP
            cursor.execute("""
                SELECT
                    cxp.Monto,
                    ISNULL(comp.MtoTotal, cxp.Monto) as MtoTotal,
                    ISNULL(comp.Contado, 0) as Contado,
                    ISNULL(comp.MtoPagos, 0) as MtoPagos,
                    cxp.NroUnico
                FROM EnterpriseAdmin_AMC.dbo.SAACXP cxp
                LEFT JOIN EnterpriseAdmin_AMC.dbo.SACOMP comp ON cxp.NumeroD = comp.NumeroD AND cxp.CodProv = comp.CodProv
                WHERE cxp.CodProv = ? AND cxp.NumeroD = ? AND cxp.TipoCxP = '10'
            """, (codprov, fac_num))
            inv_data = cursor.fetchone()
            if not inv_data:
                continue
                
            monto_pagos = float(inv_data[3])
            monto_total = float(inv_data[1])
            surplus = monto_pagos - monto_total
            
            # If there's surplus to convert to ND
            if surplus > 0.01:
                # 2. Get MAX NroUnico for Saint
                cursor.execute("SELECT ISNULL(MAX(NroUnico), 0) FROM EnterpriseAdmin_AMC.dbo.SAACXP")
                new_nro_unico_cxp = int(cursor.fetchone()[0]) + 1
                
                # 3. Create ND in SAACXP natively (TipoCxP '20')
                cursor.execute("""
                    INSERT INTO EnterpriseAdmin_AMC.dbo.SAACXP (
                        CodSucu, CodProv, NroUnico, NroRegi, FechaI, FechaE, FechaT, FechaR, FechaV,
                        CodEsta, CodUsua, CodOper, NumeroD, NumeroN, Monto, Saldo, TipoCxP,
                        CancelC, CancelT, CancelG, CancelD, EsUnPago, EsReten, DetalChq, AfectaCom, Descrip, CodTarj)
                    VALUES (
                        '00000', ?, ?, 0, GETDATE(), GETDATE(), GETDATE(), GETDATE(), GETDATE(),
                        '0', 'API', 'API', ?, ?, ?, 0, '20',
                        ?, 0, 0, 0, 0, 0, '', 0, 'Nota de Debito por Indexacion Auto', '')
                """, (codprov, new_nro_unico_cxp, nd_num, nd_ctrl, surplus, surplus))
                
                # 4. Create in SACOMP naturally (TipoCom 'J')
                cursor.execute("SELECT ISNULL(MAX(NroUnico), 0) FROM EnterpriseAdmin_AMC.dbo.SACOMP")
                new_nro_unico_comp = int(cursor.fetchone()[0]) + 1
                
                cursor.execute("""
                    INSERT INTO EnterpriseAdmin_AMC.dbo.SACOMP (
                        CodSucu, TipoCom, NumeroD, CodProv, NroUnico, NroCtrol, CodEsta, CodUsua, Signo,
                        FechaT, FechaI, FechaE, FechaV, MtoTotal, Contado, MtoPagos, Descrip, Factor, CodTarj)
                    VALUES (
                        '00000', 'J', ?, ?, ?, ?, '0', 'API', 1,
                        GETDATE(), GETDATE(), GETDATE(), GETDATE(), ?, 0, ?, 'ND Indexacion Auto', 1, '')
                """, (nd_num, codprov, new_nro_unico_comp, nd_ctrl, surplus, surplus))
                
                # 5. Restore Invoice Balance to 0 
                # (Subtract surplus from MtoPagos so MtoTotal - MtoPagos = 0)
                sanitized_mto_pagos = monto_total # It paid exactly the original amount
                
                cursor.execute("""
                    UPDATE EnterpriseAdmin_AMC.dbo.SAACXP 
                    SET Saldo = 0, CancelC = (CancelC - ?)
                    WHERE CodProv = ? AND NumeroD = ? AND TipoCxP = '10'
                """, (surplus, codprov, fac_num))
                
                cursor.execute("""
                    UPDATE EnterpriseAdmin_AMC.dbo.SACOMP 
                    SET MtoPagos = ?
                    WHERE CodProv = ? AND NumeroD = ? AND TipoCom = 'H'
                """, (sanitized_mto_pagos, codprov, fac_num))
                
                # 6. Insert Local Custom Table explicitly identifying ND creation
                cursor.execute("SELECT ISNULL(MAX(AbonoID), 0) FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos WITH (UPDLOCK)")
                new_abono_id = cursor.fetchone()[0] + 1
                
                cursor.execute("""
                    INSERT INTO EnterpriseAdmin_AMC.dbo.CxP_Abonos (
                        AbonoID, NumeroD, CodProv, FechaAbono, MontoBsAbonado, 
                        MontoUsdAbonado, TasaCambioDiaAbono, AplicaIndexacion, 
                        Referencia, FechaRegistro, RutaComprobante, NotificarCorreo, TipoAbono
                    ) VALUES (
                        ?, ?, ?, GETDATE(), ?, 0, 1, 0, ?, GETDATE(), '', 0, 'NOTA_DEBITO'
                    )
                """, (new_abono_id, fac_num, codprov, surplus, nd_num))
            
            # 7. Maintain current DebitNotesTracking table for frontend visibility exactly as it was
            cursor.execute("SELECT CodProv FROM EnterpriseAdmin_AMC.Procurement.DebitNotesTracking WHERE CodProv=? AND NumeroD=?", (inv.CodProv, inv.NumeroD))
            if cursor.fetchone():
                cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.DebitNotesTracking SET Estatus='EMITIDA', NotaDebitoID=?, MontoRetencionBs=?, FechaEmision=GETDATE() WHERE CodProv=? AND NumeroD=?", (payload.NotaDebitoID, inv.MontoRetencionBs, inv.CodProv, inv.NumeroD))
            else:
                cursor.execute("INSERT INTO EnterpriseAdmin_AMC.Procurement.DebitNotesTracking (CodProv, NumeroD, Estatus, NotaDebitoID, MontoRetencionBs, FechaEmision) VALUES (?, ?, 'EMITIDA', ?, ?, GETDATE())", (inv.CodProv, inv.NumeroD, payload.NotaDebitoID, inv.MontoRetencionBs))
        
        conn.commit()
        return {"message": "Notas de débito registradas y conciliadas correctamente en Saint."}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error register_debit_note: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.patch("/api/procurement/debit-notes/{id_nd}")
async def anular_debit_note(id_nd: int):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT Estatus, CodProv, NumeroD, NotaDebitoID FROM EnterpriseAdmin_AMC.Procurement.DebitNotesTracking WHERE Id = ?", (id_nd,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Nota de Débito no encontrada")
            
        if row[0] == 'ANULADA':
            return {"message": "La nota de débito ya se encuentra anulada."}
            
        codprov = row[1]
        fac_num = row[2]
        nd_num = row[3]
        
        # 1. Look for the ND natively to see how much we must reverse
        cursor.execute("""
            SELECT MtoTotal FROM EnterpriseAdmin_AMC.dbo.SACOMP 
            WHERE NumeroD = ? AND CodProv = ? AND TipoCom = 'J'
        """, (nd_num, codprov))
        nd_data = cursor.fetchone()
        
        if nd_data:
            surplus = float(nd_data[0])
            if surplus > 0:
                # 2. Reverse Invoice balance (restore its negative behavior)
                cursor.execute("""
                    UPDATE EnterpriseAdmin_AMC.dbo.SAACXP 
                    SET CancelC = (CancelC + ?), Saldo = (Monto - (CancelC + ?))
                    WHERE CodProv = ? AND NumeroD = ? AND TipoCxP = '10'
                """, (surplus, surplus, codprov, fac_num))
                
                cursor.execute("""
                    UPDATE EnterpriseAdmin_AMC.dbo.SACOMP 
                    SET MtoPagos = (MtoPagos + ?)
                    WHERE CodProv = ? AND NumeroD = ? AND TipoCom = 'H'
                """, (surplus, codprov, fac_num))
                
                # 3. Anular ND natively in Saint by zeroing it
                cursor.execute("""
                    UPDATE EnterpriseAdmin_AMC.dbo.SAACXP 
                    SET Monto = 0, Saldo = 0, CancelC = 0, Descrip = '**ANULADA WEB**'
                    WHERE NumeroD = ? AND CodProv = ? AND TipoCxP = '20'
                """, (nd_num, codprov))
                
                cursor.execute("""
                    UPDATE EnterpriseAdmin_AMC.dbo.SACOMP 
                    SET MtoTotal = 0, MtoPagos = 0, Descrip = '**ANULADA WEB**'
                    WHERE NumeroD = ? AND CodProv = ? AND TipoCom = 'J'
                """, (nd_num, codprov))
                
                # 4. Remove local tracking flag
                cursor.execute("""
                    DELETE FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos 
                    WHERE Referencia = ? AND CodProv = ? AND TipoAbono = 'NOTA_DEBITO'
                """, (nd_num, codprov))

        cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.DebitNotesTracking SET Estatus = 'ANULADA' WHERE Id = ?", (id_nd,))
        conn.commit()
        return {"message": "Nota de Débito anulada exitosamente y restaurado estado central."}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error anular_debit_note: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.get("/api/exchange-rate")
async def get_exchange_rate(fecha: str = None):
    try:
        logging.info(f"API: get_exchange_rate requested for fecha={fecha}")
        conn = database.get_db_connection()
        cursor = conn.cursor()
        if fecha:
            # Use a safer comparison: strictly before the next day
            query = """
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday 
                WHERE fecha < DATEADD(day, 1, CAST(? AS DATE))
                AND dolarbcv IS NOT NULL
                ORDER BY fecha DESC
            """
            cursor.execute(query, (fecha,))
        else:
            query = """
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday 
                WHERE dolarbcv IS NOT NULL
                ORDER BY fecha DESC
            """
            cursor.execute(query)
            
        row = cursor.fetchone()
        if not row:
            logging.warning(f"API: No rate found for fecha={fecha}")
            return {"rate": None}
            
        rate = float(row[0])
        logging.info(f"API: Returning rate={rate} for fecha={fecha}")
        return {"rate": rate}
    except Exception as e:
        logging.error(f"Error fetching exchange rate: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.get("/api/reports/aging")
async def report_aging():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        query = """
            SELECT
              SAPROV.CodProv,
              SAPROV.Descrip AS Proveedor,
              SUM(CASE WHEN DATEDIFF(day, SAACXP.FechaV, GETDATE()) <= 0 THEN SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0) ELSE 0 END) AS PorVencer,
              SUM(CASE WHEN DATEDIFF(day, SAACXP.FechaV, GETDATE()) BETWEEN 1 AND 30 THEN SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0) ELSE 0 END) AS Dias_1_30,
              SUM(CASE WHEN DATEDIFF(day, SAACXP.FechaV, GETDATE()) BETWEEN 31 AND 60 THEN SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0) ELSE 0 END) AS Dias_31_60,
              SUM(CASE WHEN DATEDIFF(day, SAACXP.FechaV, GETDATE()) BETWEEN 61 AND 90 THEN SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0) ELSE 0 END) AS Dias_61_90,
              SUM(CASE WHEN DATEDIFF(day, SAACXP.FechaV, GETDATE()) > 90 THEN SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0) ELSE 0 END) AS Mas_90,
              SUM(SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0)) AS Total
            FROM dbo.SAACXP WITH (NOLOCK)
            LEFT JOIN dbo.SAPROV WITH (NOLOCK) ON SAACXP.CodProv = SAPROV.CodProv
            OUTER APPLY (
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday WITH (NOLOCK)
                WHERE CAST(fecha AS DATE) <= CAST(SAACXP.FechaE AS DATE)
                ORDER BY fecha DESC
            ) dt_emision
            WHERE SAACXP.Saldo > 0 AND SAACXP.TipoCxP = '10'
            GROUP BY SAPROV.CodProv, SAPROV.Descrip
            ORDER BY Total DESC
        """
        cursor.execute(query)
        columns = [column[0] for column in cursor.description]
        data = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Ensure floating point for JSON serialization
        for item in data:
            for k in ['PorVencer', 'Dias_1_30', 'Dias_31_60', 'Dias_61_90', 'Mas_90', 'Total']:
                item[k] = float(item[k]) if item[k] else 0
                
        return {"data": data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.get("/api/reports/cashflow")
async def report_cashflow(desde: str = None, hasta: str = None):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        date_filter_pagos = ""
        date_filter_gastos = ""
        params_pagos = []
        params_gastos = []

        if desde:
            date_filter_pagos += "CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE) >= ?"
            date_filter_gastos += "CAST(fecha_proyectada AS DATE) >= ?"
            params_pagos.append(desde)
            params_gastos.append(desde)
            
        if hasta:
            if desde:
                date_filter_pagos += " AND "
                date_filter_gastos += " AND "
            date_filter_pagos += "CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE) <= ?"
            date_filter_gastos += "CAST(fecha_proyectada AS DATE) <= ?"
            params_pagos.append(hasta)
            params_gastos.append(hasta)
            
        if not date_filter_pagos:
            date_filter_pagos = "1=1"
        if not date_filter_gastos:
            date_filter_gastos = "1=1"

        params = params_pagos + params_gastos + params_gastos

        query = f"""
            WITH Facturas AS (
                SELECT 
                   CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE) AS Fecha,
                   SUM(SAACXP.Saldo) AS SaldoProyectado,
                   SUM(SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0)) AS SaldoProyectadoUSD
                FROM dbo.SAACXP WITH (NOLOCK)
                LEFT JOIN EnterpriseAdmin_AMC.Procurement.PagosPlanificados PP ON SAACXP.NroUnico = PP.NroUnico
                OUTER APPLY (
                    SELECT TOP 1 dolarbcv 
                    FROM EnterpriseAdmin_AMC.dbo.dolartoday 
                    WHERE CAST(fecha AS DATE) <= CAST(SAACXP.FechaE AS DATE)
                    ORDER BY fecha DESC
                ) dt_emision
                WHERE SAACXP.Saldo > 0 AND SAACXP.TipoCxP = '10' AND {date_filter_pagos}
                GROUP BY CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE)
            ),
            GastosFijos AS (
                SELECT 
                    CAST(fecha_proyectada AS DATE) AS Fecha,
                    SUM(monto_usd) AS SalidasFijosUSD
                FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados WITH (NOLOCK)
                WHERE tipo = 'Farmacia' AND estado = 'Pendiente' AND {date_filter_gastos}
                GROUP BY CAST(fecha_proyectada AS DATE)
            ),
            GastosPersonales AS (
                SELECT 
                    CAST(fecha_proyectada AS DATE) AS Fecha,
                    SUM(monto_usd) AS SalidasPersonalesUSD
                FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados WITH (NOLOCK)
                WHERE tipo = 'Personal' AND estado = 'Pendiente' AND {date_filter_gastos}
                GROUP BY CAST(fecha_proyectada AS DATE)
            ),
            Fechas AS (
                SELECT Fecha FROM Facturas
                UNION
                SELECT Fecha FROM GastosFijos
                UNION
                SELECT Fecha FROM GastosPersonales
            )
            SELECT 
                FORMAT(F.Fecha, 'yyyy-MM-dd') AS Periodo,
                CAST(ISNULL(ROUND(FCT.SaldoProyectado, 2), 0) AS FLOAT) AS SaldoProyectado,
                CAST(ISNULL(FCT.SaldoProyectadoUSD, 0) AS FLOAT) AS FacturasUSD,
                CAST(ISNULL(GF.SalidasFijosUSD, 0) AS FLOAT) AS GastosFijosUSD,
                CAST(ISNULL(GP.SalidasPersonalesUSD, 0) AS FLOAT) AS GastosPersonalesUSD,
                CAST(ISNULL(FCT.SaldoProyectadoUSD, 0) + ISNULL(GF.SalidasFijosUSD, 0) + ISNULL(GP.SalidasPersonalesUSD, 0) AS FLOAT) AS SaldoProyectadoUSD
            FROM Fechas F
            LEFT JOIN Facturas FCT ON F.Fecha = FCT.Fecha
            LEFT JOIN GastosFijos GF ON F.Fecha = GF.Fecha
            LEFT JOIN GastosPersonales GP ON F.Fecha = GP.Fecha
            ORDER BY F.Fecha
        """
        cursor.execute(query, params)
        columns = [column[0] for column in cursor.description]
        return {"data": [dict(zip(columns, row)) for row in cursor.fetchall()]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()



@router.get("/api/reports/dpo")
async def report_dpo():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        query = """
            SELECT 
                FORMAT(SAACXP.FechaE, 'yyyy-MM') AS Periodo,
                AVG(DATEDIFF(day, SAACXP.FechaE, SAPAGCXP.FechaE)) AS PromedioDiasPago,
                COUNT(SAACXP.NroUnico) AS FacturasPagadas
            FROM dbo.SAACXP
            INNER JOIN dbo.SAPAGCXP ON SAACXP.NroUnico = SAPAGCXP.NroUnico
            WHERE SAACXP.Saldo <= 0 AND SAACXP.TipoCxP = '10' AND SAPAGCXP.FechaE >= DATEADD(year, -1, GETDATE())
            GROUP BY FORMAT(SAACXP.FechaE, 'yyyy-MM')
            ORDER BY Periodo
        """
        cursor.execute(query)
        columns = [column[0] for column in cursor.description]
        return {"data": [dict(zip(columns, row)) for row in cursor.fetchall()]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

# --- FORECAST & EVENTS ENDPOINTS ---

@router.get("/api/reports/forecast-sales")
async def report_forecast_sales(desde: str = None, hasta: str = None):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        where_ventas = ""
        where_reales = ""
        params_ventas = []
        params_reales = []
        
        if desde:
            where_ventas += " AND CAST(fecha_proyeccion AS DATE) >= ?"
            where_reales += " AND CAST(fecha AS DATE) >= ?"
            params_ventas.append(desde)
            params_reales.append(desde)
        else:
            where_ventas += " AND CAST(fecha_proyeccion AS DATE) >= CAST(GETDATE() AS DATE)"
            where_reales += " AND CAST(fecha AS DATE) >= DATEADD(day, -30, CAST(GETDATE() AS DATE))"
            
        if hasta:
            where_ventas += " AND CAST(fecha_proyeccion AS DATE) <= ?"
            where_reales += " AND CAST(fecha AS DATE) <= ?"
            params_ventas.append(hasta)
            params_reales.append(hasta)
            
        where_reales += " AND CAST(fecha AS DATE) < CAST(GETDATE() AS DATE)"
        where_ventas += " AND CAST(fecha_proyeccion AS DATE) >= CAST(GETDATE() AS DATE)"

        query = f"""
            SELECT 
                FORMAT(CAST(fecha_proyeccion AS DATE), 'yyyy-MM-dd') AS Periodo,
                CAST(monto_proyectado_ves AS FLOAT) AS VentasProyectadas,
                CAST(monto_proyectado_usd AS FLOAT) AS VentasProyectadasUSD
            FROM EnterpriseAdmin_AMC.Procurement.sales_forecast WITH (NOLOCK)
            WHERE 1=1 {where_ventas}
            UNION ALL
            SELECT 
                FORMAT(CAST(fecha AS DATE), 'yyyy-MM-dd') AS Periodo,
                CAST(SUM(MtoVentas) AS FLOAT) AS VentasProyectadas,
                CAST(SUM(CAST(MtoDolar AS FLOAT)) AS FLOAT) AS VentasProyectadasUSD
            FROM EnterpriseAdmin_AMC.dbo.CUSTOM_SAEVTA WITH (NOLOCK)
            WHERE 1=1 {where_reales}
            GROUP BY CAST(fecha AS DATE), FORMAT(CAST(fecha AS DATE), 'yyyy-MM-dd')
            ORDER BY Periodo
        """
        params = params_ventas + params_reales
        
        cursor.execute(query, params)
        columns = [column[0] for column in cursor.description]
        return {"data": [dict(zip(columns, row)) for row in cursor.fetchall()]}
    except Exception as e:
        logging.error(f"Error in forecast-sales: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.get("/api/reports/forecast-consolidated")
async def report_forecast_consolidated(
    desde: str = None, 
    hasta: str = None,
    fecha_arranque: str = None,
    caja_usd: float = 0.0,
    caja_bs: float = 0.0,
    delay_days: int = 1
):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # If no start date provided, use today
        if not fecha_arranque:
            from datetime import date
            fecha_arranque = date.today().isoformat()
            
        p_pagos = [fecha_arranque]
        p_ventas = [fecha_arranque]
        p_ventas_real = [fecha_arranque]
        p_gastos = [fecha_arranque]

        # Everything is calculated ALWAYS from fecha_arranque (Day Zero).
        # We fetch all history from Day Zero to compute the running total properly.
        # The frontend will just "hide" rows using JS if `desde` > `fecha_arranque`, 
        # but the backend must compute everything from Day Zero.
        date_filter_pagos = "CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE) >= ?"
        date_filter_ventas = "CAST(DATEADD(day, ?, fecha_proyeccion) AS DATE) > ?"
        date_filter_ventas_real = "CAST(DATEADD(day, ?, fecha) AS DATE) > ?"
        date_filter_gastos = "CAST(fecha_proyectada AS DATE) >= ?"
        
        # Insert delay_days before fecha_arranque for Ventas DateAdd
        p_ventas.insert(0, delay_days)
        p_ventas_real.insert(0, delay_days)
        
        if hasta:
            date_filter_pagos += " AND CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE) <= ?"
            date_filter_ventas += " AND CAST(DATEADD(day, ?, fecha_proyeccion) AS DATE) <= ?"
            date_filter_ventas_real += " AND CAST(DATEADD(day, ?, fecha) AS DATE) <= ?"
            date_filter_gastos += " AND CAST(fecha_proyectada AS DATE) <= ?"
            p_pagos.append(hasta)
            p_ventas.extend([delay_days, hasta])
            p_ventas_real.extend([delay_days, hasta])
            p_gastos.append(hasta)
            
        # We only want REAL sales for past dates
        date_filter_ventas_real += " AND CAST(fecha AS DATE) < CAST(GETDATE() AS DATE)"
        # We only want FORECAST sales for today and future
        date_filter_ventas += " AND CAST(fecha_proyeccion AS DATE) >= CAST(GETDATE() AS DATE)"

        # Prepare params in exact order of CTE execution
        params = p_pagos + p_ventas + p_ventas_real + p_gastos + p_gastos

        query = f"""
            WITH Pagos AS (
                SELECT 
                    CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE) AS Fecha,
                    SUM(SAACXP.Saldo) AS SalidasBs,
                    SUM(SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0)) AS SalidasPagosUSD
                FROM dbo.SAACXP WITH (NOLOCK)
                LEFT JOIN EnterpriseAdmin_AMC.Procurement.PagosPlanificados PP WITH (NOLOCK) ON SAACXP.NroUnico = PP.NroUnico
                OUTER APPLY (
                    SELECT TOP 1 dolarbcv 
                    FROM EnterpriseAdmin_AMC.dbo.dolartoday WITH (NOLOCK)
                    WHERE CAST(fecha AS DATE) <= CAST(SAACXP.FechaE AS DATE)
                    ORDER BY fecha DESC
                ) dt_emision
                WHERE SAACXP.Saldo > 0 AND {date_filter_pagos}
                GROUP BY CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE)
            ),
            Ventas AS (
                SELECT 
                    CAST(DATEADD(day, {delay_days}, fecha_proyeccion) AS DATE) AS Fecha,
                    SUM(monto_proyectado_ves) AS EntradasBs,
                    SUM(monto_proyectado_usd) AS EntradasUSD
                FROM EnterpriseAdmin_AMC.Procurement.sales_forecast WITH (NOLOCK)
                WHERE {date_filter_ventas}
                GROUP BY CAST(DATEADD(day, {delay_days}, fecha_proyeccion) AS DATE)
                UNION ALL
                SELECT
                    CAST(DATEADD(day, {delay_days}, fecha) AS DATE) AS Fecha,
                    SUM(MtoVentas) AS EntradasBs,
                    SUM(CAST(MtoDolar AS FLOAT)) AS EntradasUSD
                FROM EnterpriseAdmin_AMC.dbo.CUSTOM_SAEVTA WITH (NOLOCK)
                WHERE {date_filter_ventas_real}
                GROUP BY CAST(DATEADD(day, {delay_days}, fecha) AS DATE)
            ),
            GastosFarmacia AS (
                SELECT 
                    CAST(fecha_proyectada AS DATE) AS Fecha,
                    SUM(monto_usd) AS SalidasFarmaciaUSD
                FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados WITH (NOLOCK)
                WHERE tipo = 'Farmacia' AND estado = 'Pendiente' AND {date_filter_gastos}
                GROUP BY CAST(fecha_proyectada AS DATE)
            ),
            GastosPersonales AS (
                SELECT 
                    CAST(fecha_proyectada AS DATE) AS Fecha,
                    SUM(monto_usd) AS SalidasPersonalesUSD
                FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados WITH (NOLOCK)
                WHERE tipo = 'Personal' AND estado = 'Pendiente' AND {date_filter_gastos}
                GROUP BY CAST(fecha_proyectada AS DATE)
            ),
            Fechas AS (
                SELECT Fecha FROM Pagos
                UNION
                SELECT Fecha FROM Ventas
                UNION
                SELECT Fecha FROM GastosFarmacia
                UNION
                SELECT Fecha FROM GastosPersonales
            ),
            ResumenDiario AS (
                SELECT 
                    FORMAT(F.Fecha, 'yyyy-MM-dd') AS Periodo,
                    F.Fecha AS RealFecha,
                    CAST(
                        CASE 
                            WHEN FORMAT(F.Fecha, 'yyyy-MM-dd') = '{fecha_arranque}' THEN 0 
                            ELSE ISNULL(ROUND(SUM(V.EntradasUSD), 2), 0) 
                        END AS FLOAT
                    ) AS EntradasUSD,
                    CAST(ISNULL(ROUND(SUM(P.SalidasPagosUSD), 2), 0) AS FLOAT) AS SalidasPagosUSD,
                    CAST(ISNULL(ROUND(SUM(GF.SalidasFarmaciaUSD), 2), 0) AS FLOAT) AS SalidasFarmaciaUSD,
                    CAST(ISNULL(ROUND(SUM(GP.SalidasPersonalesUSD), 2), 0) AS FLOAT) AS SalidasPersonalesUSD
                FROM Fechas F
                LEFT JOIN Ventas V ON F.Fecha = V.Fecha
                LEFT JOIN Pagos P ON F.Fecha = P.Fecha
                LEFT JOIN GastosFarmacia GF ON F.Fecha = GF.Fecha
                LEFT JOIN GastosPersonales GP ON F.Fecha = GP.Fecha
                GROUP BY F.Fecha
            )
            SELECT 
                Periodo,
                EntradasUSD,
                SalidasPagosUSD,
                SalidasFarmaciaUSD,
                SalidasPersonalesUSD,
                
                -- Saldo Real Acumulado usando SUM() OVER()
                CAST(ROUND(
                    -- Caja Inicial Total en USD (Caja USD + CajaBs convertido)
                    ( ? + ( ? / NULLIF((SELECT TOP 1 dolarbcv FROM EnterpriseAdmin_AMC.dbo.dolartoday WHERE CAST(fecha AS DATE) <= CAST(? AS DATE) ORDER BY fecha DESC), 0) ) )
                    + SUM(EntradasUSD - SalidasPagosUSD - SalidasFarmaciaUSD - SalidasPersonalesUSD) OVER (ORDER BY RealFecha ROWS UNBOUNDED PRECEDING)
                , 2) AS FLOAT) AS SaldoRealCajaUSD
                
            FROM ResumenDiario
            ORDER BY RealFecha
        """
        
        # Add the parameters for the Caja calculation at the end: Caja USD, Caja Bs, Fecha Arranque Tasa
        params.extend([caja_usd, caja_bs, fecha_arranque])

        cursor.execute(query, params)
        columns = [column[0] for column in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Apply strict UI date filtering since the backend calculated from Day Zero to get correct rolling totals
        filtered_results = []
        if not desde:
            desde = '1900-01-01'
            
        for row in results:
            if row['Periodo'] >= desde:
                if hasta:
                    if row['Periodo'] <= hasta:
                        filtered_results.append(row)
                else:
                    filtered_results.append(row)
            
        return {"data": filtered_results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

from pydantic import BaseModel
class AntigravityRequest(BaseModel):
    porcentaje_flujo: float = 0.90 
    caja_usd: float = 0.0
    caja_bs: float = 0.0
    fecha_arranque: str = None
    delay_days: int = 1 
    wacc: float = 0.12
    max_credit: float = 0.0

@router.post("/api/antigravity/run")
async def run_antigravity_optimizer(payload: AntigravityRequest):
    try:
        from antigravity_core import AntigravityEngine
        conn = database.get_db_connection()
        engine = AntigravityEngine(conn, annual_wacc=payload.wacc / 100.0 if payload.wacc > 1 else payload.wacc)
        cursor = conn.cursor()
        
        # 1. Fetch Invoices
        inv_query = """
            SELECT 
                SAACXP.NroUnico, SAACXP.NumeroD, SAPROV.CodProv, SAPROV.Descrip,
                SAACXP.Saldo, SAACXP.FechaE, SAACXP.FechaV,
                ISNULL(dt_emision.dolarbcv, 1) AS tc_emision,
                ISNULL(PC.DiasNoIndexacion, 15) AS t_tolerance,
                ISNULL(PC.DiasVencimiento, 30) AS t_due,
                ISNULL(PC.DescuentoBase_Pct, 0) / 100.0 AS desc_base_pct,
                ISNULL(PC.DescuentoBase_Condicion, 'INDEPENDIENTE') AS desc_base_cond
            FROM dbo.SAACXP WITH (NOLOCK)
            LEFT JOIN dbo.SAPROV WITH (NOLOCK) ON SAACXP.CodProv = SAPROV.CodProv
            LEFT JOIN EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones PC WITH (NOLOCK) ON SAPROV.CodProv = PC.CodProv
            OUTER APPLY (
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday WITH (NOLOCK)
                WHERE CAST(fecha AS DATE) <= CAST(SAACXP.FechaE AS DATE) AND dolarbcv IS NOT NULL
                ORDER BY fecha DESC
            ) dt_emision
            WHERE SAACXP.Saldo > 0 AND SAACXP.TipoCxP = '10'
        """
        cursor.execute(inv_query)
        columns = [column[0] for column in cursor.description]
        db_invoices = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Fetch dynamic discounts
        cursor.execute("SELECT CodProv, DiasDesde, DiasHasta, Porcentaje FROM EnterpriseAdmin_AMC.Procurement.ProveedorDescuentosProntoPago WITH (NOLOCK)")
        desc_cols = [c[0] for c in cursor.description]
        descuentos_db = [dict(zip(desc_cols, row)) for row in cursor.fetchall()]
        desc_map = {}
        for d in descuentos_db:
            desc_map.setdefault(d['CodProv'], []).append({
                "DiasDesde": int(d['DiasDesde']),
                "DiasHasta": int(d['DiasHasta']),
                # Convert to proportion
                "Porcentaje": float(d['Porcentaje']) / 100.0 
            })
            
        from datetime import date, timedelta
        today_date = date.today()
        
        invoices_payload = []
        for row in db_invoices:
            if isinstance(row["FechaE"], str):
                from datetime import datetime
                fecha_e = datetime.fromisoformat(row["FechaE"].split("T")[0]).date()
            else:
                fecha_e = row["FechaE"].date()
                
            if isinstance(row["FechaV"], str):
                from datetime import datetime
                fecha_v = datetime.fromisoformat(row["FechaV"].split("T")[0]).date()
            elif row["FechaV"]:
                fecha_v = row["FechaV"].date()
            else:
                # Fallback to configured days if SAACXP.FechaV is null
                fecha_v = fecha_e + timedelta(days=int(row["t_due"]))
                
            elapsed = (today_date - fecha_e).days
            calc_t_due = max(0, (fecha_v - fecha_e).days)
            
            invoices_payload.append({
                "id": str(row["NumeroD"]),
                "supplier": row["Descrip"],
                "nominal_bs": float(row["Saldo"]),
                "tc_emision": float(row["tc_emision"]),
                "days_elapsed_since_emission": elapsed,
                "t_tolerance": int(row["t_tolerance"]),
                "t_due": calc_t_due,
                "desc_base_pct": float(row["desc_base_pct"]),
                "desc_base_cond": row["desc_base_cond"],
                "descuentos": desc_map.get(row["CodProv"], []),
                "priority": "Media" 
            })
            
        d_fecha = payload.fecha_arranque if payload.fecha_arranque else today_date.isoformat()
        await_flow = await report_forecast_consolidated(
            desde=d_fecha, 
            hasta=None, 
            fecha_arranque=d_fecha,
            caja_usd=payload.caja_usd,
            caja_bs=payload.caja_bs,
            delay_days=payload.delay_days
        )
        cashflow_timeline = await_flow["data"]
        
        result = engine.optimize_payable_schedule(cashflow_timeline, invoices_payload, payload.porcentaje_flujo, max_credit=payload.max_credit)
        
        return result
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        import logging
        logging.error(f"Engine Error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()


@router.get("/api/forecast-events")
async def get_forecast_events():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, FORMAT(fecha, 'yyyy-MM-dd') as fecha, tipo_evento, valor FROM EnterpriseAdmin_AMC.Procurement.forecast_events ORDER BY fecha DESC")
        columns = [column[0] for column in cursor.description]

        return {"data": [dict(zip(columns, row)) for row in cursor.fetchall()]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.post("/api/forecast-events")
async def add_forecast_event(event: ForecastEventRequest):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        query = "INSERT INTO EnterpriseAdmin_AMC.Procurement.forecast_events (fecha, tipo_evento, valor) VALUES (?, ?, ?)"
        cursor.execute(query, (event.fecha, event.tipo_evento, event.valor))
        conn.commit()

        return {"message": "Event added successfully"}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

# --- Rutas de Plantillas de Gastos ---

@router.get("/api/expense-templates")
async def get_expense_templates():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, descripcion, tipo, monto_estimado_usd as monto_usd, dia_mes_estimado FROM EnterpriseAdmin_AMC.Procurement.PlantillasGastos ORDER BY tipo, descripcion")
        columns = [column[0] for column in cursor.description]
        return {"data": [dict(zip(columns, row)) for row in cursor.fetchall()]}
    except Exception as e:
        logging.error(f"Error in get_expense_templates: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.post("/api/expense-templates")
async def save_expense_template(template: ExpenseTemplateRequest):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        if template.id:
            query = """
                UPDATE EnterpriseAdmin_AMC.Procurement.PlantillasGastos 
                SET descripcion=?, tipo=?, monto_estimado_usd=?, dia_mes_estimado=? 
                WHERE id=?
            """
            cursor.execute(query, (template.descripcion, template.tipo, template.monto_estimado_usd, template.dia_mes_estimado, template.id))
        else:
            query = """
                INSERT INTO EnterpriseAdmin_AMC.Procurement.PlantillasGastos 
                (descripcion, tipo, monto_estimado_usd, dia_mes_estimado) 
                VALUES (?, ?, ?, ?)
            """
            cursor.execute(query, (template.descripcion, template.tipo, template.monto_estimado_usd, template.dia_mes_estimado))
            
        conn.commit()
        return {"message": "Plantilla guardada exitosamente"}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.delete("/api/expense-templates/{id}")
async def delete_expense_template(id: int):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM EnterpriseAdmin_AMC.Procurement.PlantillasGastos WHERE id = ?", (id,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Plantilla no encontrada")
        conn.commit()
        return {"message": "Plantilla eliminada"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

# --- Rutas de Modulo Batch/Generación Mensual ---

@router.get("/api/expenses/generate-batch/{mes}/{anio}")
async def get_expense_batch(mes: int, anio: int):
    """
    Simula qué debería pagarse en base a las plantillas y el mes para que la UI lo verifique.
    """
    try:
        from datetime import date
        pass  # Just ensuring it works locally
    except:
        pass
    
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, descripcion, tipo, monto_estimado_usd, dia_mes_estimado FROM EnterpriseAdmin_AMC.Procurement.PlantillasGastos")
        
        batch = []
        for row in cursor.fetchall():
            dia = row.dia_mes_estimado
            # Limitar dia al máximo de días del mes si es un día como 31 y el mes es febrero
            try:
                import calendar
                max_dia = calendar.monthrange(anio, mes)[1]
                dia = min(dia, max_dia)
                fecha_proy = f"{anio}-{mes:02d}-{dia:02d}"
            except:
                fecha_proy = f"{anio}-{mes:02d}-15" # Fallback
                
            batch.append({
                "template_id": row.id,
                "descripcion": row.descripcion,
                "tipo": row.tipo,
                "monto_usd": row.monto_estimado_usd,
                "fecha_proyectada": fecha_proy,
                "estado": "Pendiente"
            })
            
        return {"data": batch}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.post("/api/expenses/batch")
async def save_expense_batch(payload: BatchExpenseRequest):
    """
    Inserta o limpia el mes y guarda los gastos confirmados desde la interfaz.
    """
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Elimina SOLO los registros cuya descripcion viene en el payload (los que el usuario seleccionó)
        # Si no se envían descripciones, borra todo el mes (comportamiento legacy)
        if payload.descripcionesAEliminar:
            for desc in payload.descripcionesAEliminar:
                cursor.execute(
                    "DELETE FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados "
                    "WHERE YEAR(fecha_proyectada) = ? AND MONTH(fecha_proyectada) = ? "
                    "AND (is_adhoc = 0 OR is_adhoc IS NULL) AND descripcion = ?",
                    (payload.anio, payload.mes, desc)
                )
        else:
            cursor.execute(
                "DELETE FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados "
                "WHERE YEAR(fecha_proyectada) = ? AND MONTH(fecha_proyectada) = ? "
                "AND (is_adhoc = 0 OR is_adhoc IS NULL)",
                (payload.anio, payload.mes)
            )
        
        # Insertar los nuevos generados de la UI
        for e in payload.gastos:
            cursor.execute(
                "INSERT INTO EnterpriseAdmin_AMC.Procurement.GastosProgramados (descripcion, tipo, monto_usd, fecha_proyectada, estado, is_adhoc) VALUES (?, ?, ?, ?, ?, 0)",
                (e.descripcion, e.tipo, e.monto_usd, e.fecha_proyectada, e.estado)
            )
            
        conn.commit()
        return {"message": "Lote mensual guardado exitosamente"}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.post("/api/expenses/programmed/single")
async def save_single_expense(expense: ProgrammedExpense):
    """
    Guarda un gasto variable on-the-fly directamente a la BD y lo marca como adhoc.
    """
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO EnterpriseAdmin_AMC.Procurement.GastosProgramados (descripcion, tipo, monto_usd, fecha_proyectada, estado, is_adhoc) VALUES (?, ?, ?, ?, ?, 1)",
            (expense.descripcion, expense.tipo, expense.monto_usd, expense.fecha_proyectada, expense.estado)
        )
        conn.commit()
        return {"message": "Gasto ad-hoc guardado exitosamente"}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()
        
@router.get("/api/expenses/programmed")
async def get_programmed_expenses(mes: int = None, anio: int = None):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        where_clause = ""
        params = []
        if mes and anio:
             where_clause = "WHERE YEAR(fecha_proyectada) = ? AND MONTH(fecha_proyectada) = ?"
             params.extend([anio, mes])
        
        query = f"SELECT id, descripcion, tipo, monto_usd, FORMAT(fecha_proyectada, 'yyyy-MM-dd') as fecha_proyectada, estado FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados {where_clause} ORDER BY fecha_proyectada ASC"
        cursor.execute(query, tuple(params))
        columns = [column[0] for column in cursor.description]
        return {"data": [dict(zip(columns, row)) for row in cursor.fetchall()]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()
        
@router.delete("/api/expenses/programmed/{id}")
async def delete_programmed_expense(id: int):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados WHERE id = ?", (id,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Gasto no encontrado")
        conn.commit()
        return {"message": "Gasto eliminado"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.patch("/api/expenses/programmed/{id}")
async def patch_programmed_expense(id: int, payload: dict = Body(...)):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        if "fecha_proyectada" in payload:
            cursor.execute(
                "UPDATE EnterpriseAdmin_AMC.Procurement.GastosProgramados SET fecha_proyectada = ? WHERE id = ?",
                (payload["fecha_proyectada"], id)
            )
            if cursor.rowcount == 0:
                raise HTTPException(status_code=404, detail="Gasto no encontrado")
            conn.commit()
            return {"message": "Fecha actualizada"}
        raise HTTPException(status_code=400, detail="Campo no soportado")
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.delete("/api/forecast-events/{event_id}")
async def delete_forecast_event(event_id: int):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        query = "DELETE FROM EnterpriseAdmin_AMC.Procurement.forecast_events WHERE id = ?"
        cursor.execute(query, (event_id,))

        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Event not found")
        conn.commit()
        return {"message": "Event deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.get("/api/export/{report_type}")
async def export_xlsx(report_type: str, desde: Optional[str] = None, hasta: Optional[str] = None):
    try:
        data = []
        filename = f"{report_type}_reporte.xlsx"

        if report_type == "cuentas-por-pagar":
            res = await get_cuentas_por_pagar("", desde, hasta)
            data = res.get("data", [])
        elif report_type == "aging":
            res = await report_aging()
            data = res.get("data", [])
        elif report_type == "compras":
            res = await report_compras(desde, hasta)
            data = res.get("data", [])
        elif report_type == "debit-notes":
            res = await get_debit_notes()
            data = res.get("data", [])
        else:
            raise HTTPException(status_code=400, detail="Invalid report type")

        if not data:
            raise HTTPException(status_code=404, detail="No data to export")

        import openpyxl  # type: ignore
        from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type[:31]

        # Header row with styling
        headers = list(data[0].keys())
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, key in enumerate(headers, start=1):
                val = row_data.get(key)
                # Convert date/datetime to string for readability
                if hasattr(val, 'isoformat'):
                    val = str(val)[0:19]  # type: ignore
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(int(max_len) + 3, 40)

        # Save to bytes buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error exporting XLSX: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ==========================================
# RETENCIONES DE IVA MODULE
# ==========================================

@router.get("/api/retenciones/config")
async def get_retenciones_config():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT RIF_Agente, RazonSocial_Agente, DireccionFiscal_Agente, ValorUT, UltimoSecuencial, TasaEmisionSource, MontoUsdSource FROM EnterpriseAdmin_AMC.Procurement.Retenciones_Config WHERE Id = 1")
        row = cursor.fetchone()
        if not row:
            return {"data": {}}
        
        data = {
            "RifAgente": row[0],
            "NombreAgente": row[1],
            "DireccionAgente": row[2],
            "ValorUT": float(row[3]) if row[3] else 0,
            "ProximoSecuencial": row[4],
            "TasaEmisionSource": row[5] or "SACOMP",
            "MontoUsdSource": row[6] or "Calculado"
        }
        return {"data": data}
    except Exception as e:
        logging.error(f"Error checking config: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.put("/api/retenciones/config")
async def update_retenciones_config(payload: dict = Body(...)):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE EnterpriseAdmin_AMC.Procurement.Retenciones_Config 
            SET RIF_Agente = ?, RazonSocial_Agente = ?, DireccionFiscal_Agente = ?, ValorUT = ?, UltimoSecuencial = ?, TasaEmisionSource = ?, MontoUsdSource = ?
            WHERE Id = 1
        """, (
            payload.get("RifAgente"), 
            payload.get("NombreAgente"), 
            payload.get("DireccionAgente"), 
            float(payload.get("ValorUT", 0)),
            int(payload.get("ProximoSecuencial", 0)),
            payload.get("TasaEmisionSource", "SACOMP"),
            payload.get("MontoUsdSource", "Calculado")
        ))
        conn.commit()
        return {"message": "Configuración actualizada"}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error updating retenciones config: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.get("/api/retenciones")
async def get_retenciones(desde: Optional[str] = None, hasta: Optional[str] = None):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        query = """
            SELECT r.*, p.Descrip AS ProveedorNombre 
            FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA r
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON r.CodProv = p.CodProv
            WHERE 1=1
        """
        params = []
        if desde:
            query += " AND FechaRetencion >= ?"
            params.append(desde + " 00:00:00")
        if hasta:
            query += " AND FechaRetencion <= ?"
            params.append(hasta + " 23:59:59")
            
        query += " ORDER BY FechaRetencion DESC, Id DESC"
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        columns = [column[0] for column in cursor.description]
        data = [dict(zip(columns, row)) for row in rows]
        for item in data:
            if item.get('ProveedorNombre'):
                item['CodProv'] = f"{item['CodProv']} - {item['ProveedorNombre']}"
        return {"data": data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.post("/api/retenciones")
async def crear_retencion(payload: dict = Body(...)):
    """Create retention records. Supports batch: {FechaRetencion, facturas: [{NumeroD, CodProv, ...}, ...]}"""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        from datetime import datetime
        now = datetime.now()

        # Support batch (new format) or single (legacy format)
        facturas = payload.get("facturas", [payload])  # If no facturas key, treat payload as single invoice
        fecha_retencion = payload.get("FechaRetencion", now.strftime('%Y-%m-%d'))
        
        # 0. Idempotency check 
        for f in facturas:
            cursor.execute("SELECT Id FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA WITH (NOLOCK) WHERE NumeroD = ? AND CodProv = ? AND Estado != 'ANULADO'", (f["NumeroD"], f["CodProv"]))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail=f"La factura {f['NumeroD']} ya posee una retención IVA activa.")

        # 1. Generate sequential number
        cursor.execute("SELECT UltimoSecuencial FROM EnterpriseAdmin_AMC.Procurement.Retenciones_Config WITH (UPDLOCK) WHERE Id = 1")
        last_seq = cursor.fetchone()[0]
        new_seq = last_seq + 1
        
        nro_comprobante = f"{now.strftime('%Y%m')}{str(new_seq).zfill(8)}"
        
        inserted_ids = []
        for f in facturas:
            cursor.execute("""
                INSERT INTO EnterpriseAdmin_AMC.Procurement.Retenciones_IVA 
                (NumeroComprobante, NumeroD, CodProv, FechaFactura, FechaRetencion, NroControl, 
                 MontoTotal, BaseImponible, MontoExento, Alicuota, IVACausado, PorcentajeRetencion, 
                 MontoRetenido, Estado, TipoOperacion, TipoDocumento, DocAfectado)
                OUTPUT INSERTED.Id
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'EMITIDO', '01', '01', NULL)
            """, (
                nro_comprobante, f["NumeroD"], f["CodProv"], 
                f.get("FechaFactura", fecha_retencion),
                fecha_retencion, f.get("NroControl", "00-000000"),
                float(f.get("MontoTotal", 0)), float(f.get("BaseImponible", 0)), 
                float(f.get("MontoExento", 0)),
                float(f.get("Alicuota", 16)), float(f.get("IVACausado", 0)), 
                float(f.get("PorcentajeRetencion", 75)),
                float(f.get("MontoRetenido", 0))
            ))
            new_id = cursor.fetchone()[0]
            inserted_ids.append(new_id)
            
            # Auto-register as abono in CxP_Abonos with TipoAbono='RETENCION_IVA' (even if 0)
            monto_retenido = float(f.get("MontoRetenido", 0))
            if True:
                # Phase 8: Lookup mirror fields for independence
                cursor.execute("SELECT Factor, MontoMEx FROM dbo.SACOMP WITH (NOLOCK) WHERE NumeroD = ? AND CodProv = ? AND TipoCom = 'H'", (f['NumeroD'], f['CodProv']))
                sacomp_row = cursor.fetchone()
                tasa_orig = float(sacomp_row[0]) if sacomp_row and sacomp_row[0] else None
                mto_orig = float(sacomp_row[1]) if sacomp_row and sacomp_row[1] else None

                monto_usd_abonado = round(monto_retenido / tasa_orig, 4) if tasa_orig and tasa_orig > 0 else 0

                cursor.execute("SELECT ISNULL(MAX(AbonoID), 0) FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos WITH (UPDLOCK)")
                new_abono_id = cursor.fetchone()[0] + 1

                cursor.execute("""
                    INSERT INTO EnterpriseAdmin_AMC.dbo.CxP_Abonos 
                    (AbonoID, NumeroD, CodProv, FechaAbono, MontoBsAbonado, TasaCambioDiaAbono, MontoUsdAbonado, AplicaIndexacion, Referencia, TipoAbono, TasaCambioOrig, MontoMExOrig, RutaComprobante, NotificarCorreo)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    new_abono_id, f["NumeroD"], f["CodProv"], fecha_retencion,
                    monto_retenido, tasa_orig or 0, monto_usd_abonado, 0,
                    f"Retención IVA Comp. {nro_comprobante}", 'RETENCION_IVA',
                    tasa_orig, mto_orig, '', 0
                ))
        
        # 3. Update sequential
        cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.Retenciones_Config SET UltimoSecuencial = ? WHERE Id = 1", (new_seq,))
        
        conn.commit()
        logging.info(f"Retención {nro_comprobante} creada con {len(facturas)} factura(s) + abonos automáticos")
        return {"message": "Retención creada exitosamente", "NumeroComprobante": nro_comprobante, "Ids": inserted_ids}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error creating retencion: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.patch("/api/retenciones/{id_ret}")
async def anular_retencion(id_ret: int):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT Estado, NumeroComprobante, NumeroD, CodProv FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA WHERE Id = ?", (id_ret,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Retención no encontrada")
        if row[0] == 'ENTERADO':
            raise HTTPException(status_code=400, detail="No se puede anular una retención ENTERADA (declarada ante SENIAT)")
            
        cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.Retenciones_IVA SET Estado = 'ANULADO' WHERE Id = ?", (id_ret,))
        
        # Delete the auto-registered abono for this retención
        nro_comprobante = row[1]
        cursor.execute("""
            DELETE FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos 
            WHERE NumeroD = ? AND CodProv = ? AND TipoAbono = 'RETENCION_IVA' 
            AND Referencia LIKE ?
        """, (row[2], row[3], f"%{nro_comprobante}%"))
        
        conn.commit()
        return {"message": "Retención anulada y abono correspondiente eliminado"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

# --- PDF Generation for Retenciones ---
def generar_pdf_retencion(config: dict, retenciones: list) -> bytes:
    """Generate a PDF comprobante de retención IVA. Supports multiple invoices in one comprobante."""
    from fpdf import FPDF
    from datetime import datetime

    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    # Colors
    header_bg = (0, 51, 102)   # Dark blue
    header_fg = (255, 255, 255) # White
    row_alt = (230, 240, 250)  # Light blue
    
    # --- Header ---
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_fill_color(*header_bg)
    pdf.set_text_color(*header_fg)
    pdf.cell(0, 10, 'COMPROBANTE DE RETENCIÓN DEL IVA', 0, 1, 'C', fill=True)
    pdf.ln(3)
    
    # --- Agente info (Explicit Layout to prevent wrapping issues) ---
    nro_comprobante = retenciones[0]["NumeroComprobante"] if retenciones else "N/A"
    fecha_ret = str(retenciones[0].get("FechaRetencion", ""))[:10] if retenciones else "N/A"

    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "Agente de Retención:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, str(config.get("RazonSocial_Agente", "")), 0, 1)

    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "RIF:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, str(config.get("RIF_Agente", "")), 0, 1)

    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "Dirección Fiscal:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    # Using fixed width for multi_cell to guarantee clean line wrap for subsequent labels
    pdf.multi_cell(0, 6, str(config.get("DireccionFiscal_Agente", "")), 0, 'L')
    pdf.set_x(15)

    
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "Nro. Comprobante:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, str(nro_comprobante), 0, 1)

    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "Fecha de Retención:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, str(fecha_ret), 0, 1)
    
    pdf.ln(3)
    
    # --- Sujeto Retenido (from first record) ---
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "Sujeto Retenido:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    prov_name = retenciones[0].get("ProveedorNombre") or retenciones[0].get("CodProv", "")
    pdf.cell(0, 6, str(prov_name), 0, 1)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "RIF Proveedor:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, str(retenciones[0].get("CodProv", "")), 0, 1)
    pdf.ln(5)
    
    # --- Table Header ---
    # Adjusted widths to fit NEW column "Sin Derecho" (Compressed portrait layout)
    cols = [
        ("Factura", 18), ("Nro Control", 20), ("Fecha Factura", 16),
        ("Monto Total", 20), ("Sin Derecho", 20), ("Base Imponible", 20), 
        ("IVA %", 8), ("IVA Causado", 20), ("Ret. %", 8), ("Monto Retenido", 20)
    ]
    total_table_width = sum(w for n, w in cols) # ~170mm
    
    pdf.set_font('Helvetica', 'B', 6) # Slightly smaller font for more columns
    pdf.set_fill_color(*header_bg)
    pdf.set_text_color(*header_fg)
    for name, width in cols:
        pdf.cell(width, 7, name, 1, 0, 'C', fill=True)
    pdf.ln()
    
    # --- Table Rows ---
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 7)
    
    total_monto = 0
    total_exento = 0
    total_base = 0
    total_iva = 0
    total_retenido = 0
    
    for i, r in enumerate(retenciones):
        if i % 2 == 1:
            pdf.set_fill_color(*row_alt)
            fill = True
        else:
            fill = False
            
        monto = float(r.get("MontoTotal", 0))
        exento = float(r.get("MontoExento", 0))
        base = float(r.get("BaseImponible", 0))
        alicuota = float(r.get("Alicuota", 0))
        iva = float(r.get("IVACausado", base * alicuota / 100))
        pct_ret = float(r.get("PorcentajeRetencion", 0))
        retenido = float(r.get("MontoRetenido", 0))
        
        total_monto += monto
        total_exento += exento
        total_base += base
        total_iva += iva
        total_retenido += retenido
        
        fecha_fact = str(r.get("FechaFactura", ""))[:10]
        
        row_data = [
            str(r.get("NumeroD", "")),
            str(r.get("NroControl", "")),
            fecha_fact,
            f"{monto:,.2f}",
            f"{exento:,.2f}",
            f"{base:,.2f}",
            f"{alicuota:.0f}%",
            f"{iva:,.2f}",
            f"{pct_ret:.0f}%",
            f"{retenido:,.2f}"
        ]
        for j, (name, width) in enumerate(cols):
            align = 'R' if j >= 3 else 'L'
            pdf.cell(width, 6, row_data[j], 1, 0, align, fill=fill)
        pdf.ln()
    
    # --- Totals ---
    pdf.set_font('Helvetica', 'B', 6)
    pdf.set_fill_color(220, 220, 220)
    
    # Empty cells for first 3 columns
    prep_width = cols[0][1] + cols[1][1] + cols[2][1]
    pdf.cell(prep_width, 7, 'TOTALES:', 1, 0, 'R', fill=True)
    pdf.cell(cols[3][1], 7, f"{total_monto:,.2f}", 1, 0, 'R', fill=True)
    pdf.cell(cols[4][1], 7, f"{total_exento:,.2f}", 1, 0, 'R', fill=True) 
    pdf.cell(cols[5][1], 7, f"{total_base:,.2f}", 1, 0, 'R', fill=True)
    pdf.cell(cols[6][1], 7, "", 1, 0, 'C', fill=True)
    pdf.cell(cols[7][1], 7, f"{total_iva:,.2f}", 1, 0, 'R', fill=True)
    pdf.cell(cols[8][1], 7, "", 1, 0, 'C', fill=True)
    pdf.cell(cols[9][1], 7, f"{total_retenido:,.2f}", 1, 0, 'R', fill=True)
    pdf.ln(12)
    
    # --- Footer ---
    pdf.set_font('Helvetica', '', 8)
    pdf.cell(0, 5, f"Documento generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", 0, 1, 'C')
    
    out = pdf.output(dest='S')
    return out.encode('latin-1') if isinstance(out, str) else bytes(out)

@router.get("/api/retenciones/by-invoice/{numero_d}/pdf")
def get_retencion_pdf_by_invoice(numero_d: str, cod_prov: str = Query(...)):
    cod_prov = cod_prov.split(" - ")[0].strip()
    try:
        import database
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT NumeroComprobante FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA WHERE NumeroD = ? AND CodProv = ? AND Estado <> 'ANULADO'", (numero_d, cod_prov))
        row = cursor.fetchone()
        if not row or not row[0]:
            raise HTTPException(status_code=404, detail="Retención no encontrada")
        nro_comp = row[0]
        
        cursor.execute("""
            SELECT r.*, p.Descrip as ProveedorNombre
            FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA r
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON r.CodProv = p.CodProv
            WHERE r.NumeroComprobante = ? AND r.Estado <> 'ANULADO'
        """, (nro_comp,))
        rows = cursor.fetchall()
        if not rows:
            raise HTTPException(status_code=404, detail="Detalles no encontrados")
            
        cols = [column[0] for column in cursor.description]
        ret_list = [dict(zip(cols, row)) for row in rows]
        
        cursor.execute("SELECT RIF_Agente, RazonSocial_Agente, DireccionFiscal_Agente, ValorUT, UltimoSecuencial FROM EnterpriseAdmin_AMC.Procurement.Retenciones_Config WHERE Id = 1")
        cfg_row = cursor.fetchone()
        config = dict(zip([c[0] for c in cursor.description], cfg_row)) if cfg_row else {}
        
        pdf_bytes = generar_pdf_retencion(config, ret_list)
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename=Retencion_IVA_{nro_comp}.pdf"}
        )
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.get("/api/retenciones-islr/by-invoice/{numero_d}/pdf")
def get_retencion_islr_pdf_by_invoice(numero_d: str, cod_prov: str = Query(...)):
    cod_prov = cod_prov.split(" - ")[0].strip()
    try:
        import database
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT NumeroComprobante FROM EnterpriseAdmin_AMC.Procurement.Retenciones_ISLR WHERE NumeroD = ? AND CodProv = ? AND Estado <> 'ANULADO'", (numero_d, cod_prov))
        row = cursor.fetchone()
        if not row or not row[0]:
            raise HTTPException(status_code=404, detail="Retención no encontrada")
        nro_comp = row[0]
        
        cursor.execute("""
            SELECT r.*, p.Descrip as ProveedorNombre
            FROM EnterpriseAdmin_AMC.Procurement.Retenciones_ISLR r
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON r.CodProv = p.CodProv
            WHERE r.NumeroComprobante = ? AND r.Estado <> 'ANULADO'
        """, (nro_comp,))
        rows = cursor.fetchall()
        if not rows:
            raise HTTPException(status_code=404, detail="Detalles no encontrados")
            
        cols = [column[0] for column in cursor.description]
        ret_list = [dict(zip(cols, row)) for row in rows]
        
        cursor.execute("SELECT RIF_Agente, RazonSocial_Agente, DireccionFiscal_Agente, ValorUT, UltimoSecuencial FROM EnterpriseAdmin_AMC.Procurement.Retenciones_Config WHERE Id = 1")
        cfg_row = cursor.fetchone()
        config = dict(zip([c[0] for c in cursor.description], cfg_row)) if cfg_row else {}
        
        pdf_bytes = generar_pdf_islr(config, ret_list)
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename=Retencion_ISLR_{nro_comp}.pdf"}
        )
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()


@router.get("/api/retenciones/{id_ret}/pdf")
async def get_retencion_pdf(id_ret: int):
    """Generate and return PDF preview of a retention comprobante."""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Get config
        cursor.execute("SELECT RIF_Agente, RazonSocial_Agente, DireccionFiscal_Agente, ValorUT, UltimoSecuencial FROM EnterpriseAdmin_AMC.Procurement.Retenciones_Config WHERE Id = 1")
        cfg_row = cursor.fetchone()
        config = dict(zip([c[0] for c in cursor.description], cfg_row)) if cfg_row else {}
        
        # Get retention(s) - may be grouped by NumeroComprobante
        cursor.execute("""
            SELECT r.*, p.Descrip as ProveedorNombre
            FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA r
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON r.CodProv = p.CodProv
            WHERE r.Id = ?
        """, (id_ret,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Retención no encontrada")
        
        columns = [c[0] for c in cursor.description]
        main_ret = dict(zip(columns, row))
        
        # Fetch all retentions with same NumeroComprobante (for grouped comprobantes)
        cursor.execute("""
            SELECT r.*, p.Descrip as ProveedorNombre
            FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA r
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON r.CodProv = p.CodProv
            WHERE r.NumeroComprobante = ? AND r.Estado <> 'ANULADO'
            ORDER BY r.Id
        """, (main_ret["NumeroComprobante"],))
        all_rows = cursor.fetchall()
        retenciones = [dict(zip(columns, r)) for r in all_rows]
        
        pdf_bytes = generar_pdf_retencion(config, retenciones)
        
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename=Retencion_{main_ret['NumeroComprobante']}.pdf"}
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error generating PDF: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()


@router.post("/api/retenciones/{id_ret}/send-email")
async def send_retencion_email(id_ret: int):
    """Send retention comprobante via email to the provider, with PDF + Excel attachments."""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Get config
        cursor.execute("SELECT RIF_Agente, RazonSocial_Agente, DireccionFiscal_Agente, ValorUT FROM EnterpriseAdmin_AMC.Procurement.Retenciones_Config WHERE Id = 1")
        cfg_row = cursor.fetchone()
        config = dict(zip([c[0] for c in cursor.description], cfg_row)) if cfg_row else {}
        
        # Get retention
        cursor.execute("""
            SELECT r.*, p.Descrip as ProveedorNombre
            FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA r
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON r.CodProv = p.CodProv
            WHERE r.Id = ?
        """, (id_ret,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Retención no encontrada")
        
        columns = [c[0] for c in cursor.description]
        main_ret = dict(zip(columns, row))
        
        # Fetch all grouped retentions
        cursor.execute("""
            SELECT r.*, p.Descrip as ProveedorNombre
            FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA r
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON r.CodProv = p.CodProv
            WHERE r.NumeroComprobante = ?
            ORDER BY r.Id
        """, (main_ret["NumeroComprobante"],))
        all_rows = cursor.fetchall()
        retenciones = [dict(zip(columns, r)) for r in all_rows]
        
        # Get provider email
        cursor.execute("""
            SELECT c.Email, p.Descrip 
            FROM EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones c 
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON c.CodProv = p.CodProv 
            WHERE c.CodProv = ?
        """, (main_ret["CodProv"],))
        prov = cursor.fetchone()
        
        if not prov or not prov.Email:
            return {"email_sent": False, "message": "Proveedor sin email configurado."}
        
        # Generate PDF
        pdf_bytes = generar_pdf_retencion(config, retenciones)
        
        # Generate styled Excel
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        excel_data = []
        for r in retenciones:
            excel_data.append({
                "Nro Factura": r["NumeroD"],
                "Nro Control": r.get("NroControl", ""),
                "Fecha Factura": str(r.get("FechaFactura", ""))[:10],
                "Monto Total (Bs)": float(r.get("MontoTotal", 0)),
                "Base Imponible (Bs)": float(r.get("BaseImponible", 0)),
                "Alícuota IVA (%)": float(r.get("Alicuota", 0)),
                "IVA Causado (Bs)": float(r.get("IVACausado", 0)),
                "% Retención": float(r.get("PorcentajeRetencion", 0)),
                "Monto Retenido (Bs)": float(r.get("MontoRetenido", 0))
            })
        
        df = pd.DataFrame(excel_data)
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Retención IVA')
            ws = writer.sheets['Retención IVA']
            
            # Header style
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            # Data formatting
            for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row_cells:
                    cell.border = thin_border
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
            
            # Totals row
            total_row = ws.max_row + 1
            ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
            ws.cell(row=total_row, column=1).border = thin_border
            ws.cell(row=total_row, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.border = thin_border
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                # Sum numeric columns (4,5,7,9 are numeric)
                if col_idx in [4, 5, 7, 9]:
                    col_letter = chr(64 + col_idx)
                    cell.value = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
                    cell.number_format = '#,##0.00'
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="right")
            
            # Auto-fit column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 25)
        
        excel_buffer.seek(0)
        
        # Build email
        service = get_gmail_service()
        if not service:
            return {"email_sent": False, "message": "Gmail API no disponible."}
        
        # Support multiple emails separated by ;
        emails = [e.strip() for e in prov.Email.split(";") if e.strip()]
        remitente = os.getenv("SMTP_EMAIL", "")
        
        nro_comp = main_ret["NumeroComprobante"]
        prov_nombre = prov.Descrip or "Proveedor"
        
        msg = MIMEMultipart()
        msg['From'] = remitente
        msg['To'] = ", ".join(emails)
        msg['Subject'] = f"Comprobante de Retención IVA - {nro_comp}"
        
        total_retenido = sum(float(r.get("MontoRetenido", 0)) for r in retenciones)
        
        cuerpo = f"""Estimados/as {prov_nombre},

Adjunto a este correo el comprobante de retención de IVA Nro. {nro_comp}.
Monto total retenido: Bs. {total_retenido:,.2f}

Atentamente,
El equipo de Administración."""
        msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
        
        # Attach PDF
        part_pdf = MIMEBase("application", "pdf")
        part_pdf.set_payload(pdf_bytes)
        encoders.encode_base64(part_pdf)
        part_pdf.add_header("Content-Disposition", f"attachment; filename=Retencion_{nro_comp}.pdf")
        msg.attach(part_pdf)
        
        # Attach Excel
        part_excel = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part_excel.set_payload(excel_buffer.read())
        encoders.encode_base64(part_excel)
        part_excel.add_header("Content-Disposition", f"attachment; filename=Resumen_Retencion_{nro_comp}.xlsx")
        msg.attach(part_excel)
        
        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        service.users().messages().send(userId='me', body={'raw': raw}).execute()
        logging.info(f"Correo retención {nro_comp} enviado a {', '.join(emails)}")
        
        return {"email_sent": True, "message": f"Correo enviado a {', '.join(emails)}"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error sending retention email: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()


@router.get("/api/retenciones/export-txt")
async def export_retenciones_txt(desde: Optional[str] = None, hasta: Optional[str] = None):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Get Agent RIF
        cursor.execute("SELECT RIF_Agente FROM EnterpriseAdmin_AMC.Procurement.Retenciones_Config WHERE Id = 1")
        rif_agente = cursor.fetchone()[0]
        
        query = "SELECT * FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA WHERE Estado = 'EMITIDO'"
        params = []
        if desde:
            query += " AND FechaRetencion >= ?"
            params.append(desde + " 00:00:00")
        if hasta:
            query += " AND FechaRetencion <= ?"
            params.append(hasta + " 23:59:59")
            
        query += " ORDER BY FechaRetencion ASC"
        cursor.execute(query, params)
        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        data = [dict(zip(columns, row)) for row in rows]
        
        if not data:
            raise HTTPException(status_code=404, detail="No hay retenciones para exportar en este período")
            
        lines = []
        for r in data:
            # Format: 16 columns tab separated
            # 1: RIF Agente
            # 2: Periodo (AAAAMM)
            # 3: Fecha Doc (AAAA-MM-DD)
            # 4: Tipo Operacion (01, 02, 03)
            # 5: Tipo Doc (01, 02, 03)
            # 6: RIF Prov
            # 7: Nro Factura
            # 8: Nro Control
            # 9: Monto Total
            # 10: Base Imponible
            # 11: Monto Retenido
            # 12: Doc Afectado
            # 13: Nro Comprobante
            # 14: Monto Exento
            # 15: Alicuota
            # 16: Nro Expediente
            
            periodo = str(r["FechaRetencion"])[0:7].replace("-", "") # type: ignore
            fdoc = str(r["FechaFactura"])[0:10] # type: ignore
            
            line = [
                rif_agente,
                periodo,
                fdoc,
                r["TipoOperacion"],
                r["TipoDocumento"],
                str(r["CodProv"]).replace("-", ""),  # clean RIF # type: ignore
                r["NumeroD"],
                r["NroControl"],
                f"{r['MontoTotal']:.2f}",
                f"{r['BaseImponible']:.2f}",
                f"{r['MontoRetenido']:.2f}",
                str(r["DocAfectado"] or "0"),
                str(r["NumeroComprobante"]),
                f"{r['MontoExento']:.2f}",
                f"{r['Alicuota']:.2f}",
                str(r["NroExpediente"] or "0")
            ]
            lines.append("\t".join([str(x) for x in line]))
            
            # Also mark them as ENTERADO
            cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.Retenciones_IVA SET Estado = 'ENTERADO' WHERE Id = ?", (r["Id"],))
            
        conn.commit()
        
        txt_content = "\n".join(lines)
        return StreamingResponse(
            io.BytesIO(txt_content.encode("utf-8")),
            media_type="text/plain",
            headers={"Content-Disposition": f"attachment; filename=retenciones_{periodo}.txt"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

# ═══════════════════════════════════════════════════════════
# NOTAS DE CRÉDITO MODULE
# ═══════════════════════════════════════════════════════════

@router.get("/api/procurement/credit-notes")
async def get_credit_notes(cod_prov: Optional[str] = None, estatus: Optional[str] = None):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        query = """
            SELECT cn.*, prov.Descrip AS ProveedorNombre
            FROM EnterpriseAdmin_AMC.Procurement.CreditNotesTracking cn
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV prov ON cn.CodProv = prov.CodProv
            WHERE 1=1
            -- Excluir facturas donde un ajuste ya anule la Nota de Crédito
            AND NOT EXISTS (
                SELECT 1 FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos ab2
                INNER JOIN EnterpriseAdmin_AMC.Procurement.MotivosAjuste ma ON ab2.MotivoAjusteID = ma.MotivoID
                WHERE ab2.NumeroD = cn.NumeroD AND ab2.CodProv = cn.CodProv AND ma.AnulaNotaCredito = 1
            )
        """
        params = []
        if cod_prov:
            query += " AND cn.CodProv = ?"
            params.append(cod_prov)
        if estatus:
            query += " AND cn.Estatus = ?"
            params.append(estatus)
        query += " ORDER BY cn.FechaSolicitud DESC"
        cursor.execute(query, tuple(params))
        columns = [col[0] for col in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        for r in results:
            if r.get('ProveedorNombre'):
                r['CodProv'] = f"{r['CodProv']} - {r['ProveedorNombre']}"
            for k, v in r.items():
                if hasattr(v, 'quantize'):
                    r[k] = float(v) if v is not None else 0.0
                elif hasattr(v, 'strftime'):
                    r[k] = v.strftime('%Y-%m-%d %H:%M') if v else None
        return {"data": results}
    except Exception as e:
        logging.error(f"Error fetching credit notes: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.get("/api/procurement/credit-notes/pending/{cod_prov}")
async def get_pending_credit_notes(cod_prov: str):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT cn.*, prov.Descrip AS ProveedorNombre
            FROM EnterpriseAdmin_AMC.Procurement.CreditNotesTracking cn
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV prov ON cn.CodProv = prov.CodProv
            WHERE cn.CodProv = ? AND cn.Estatus = 'PENDIENTE'
            AND NOT EXISTS (
                SELECT 1 FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos ab2
                INNER JOIN EnterpriseAdmin_AMC.Procurement.MotivosAjuste ma ON ab2.MotivoAjusteID = ma.MotivoID
                WHERE ab2.NumeroD = cn.NumeroD AND ab2.CodProv = cn.CodProv AND ma.AnulaNotaCredito = 1
            )
            ORDER BY cn.FechaSolicitud DESC
        """, (cod_prov,))
        columns = [col[0] for col in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        for r in results:
            for k, v in r.items():
                if hasattr(v, 'quantize'):
                    r[k] = float(v) if v is not None else 0.0
                elif hasattr(v, 'strftime'):
                    r[k] = v.strftime('%Y-%m-%d %H:%M') if v else None
        return {"data": results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

def generar_pdf_nc_request(nc_data: dict, prov_data: dict, email_template_body: str) -> bytes:
    from fpdf import FPDF
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    header_bg = (0, 51, 102)
    header_fg = (255, 255, 255)
    
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_fill_color(*header_bg)
    pdf.set_text_color(*header_fg)
    pdf.cell(0, 10, 'SOLICITUD DE NOTA DE CREDITO', 0, 1, 'C', fill=True)
    pdf.ln(5)
    
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(0, 8, "Datos del Proveedor:", 0, 1)
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(40, 6, "Razon Social:", 0, 0)
    descrip = str(prov_data.get('Descrip', '')).encode('latin-1', 'replace').decode('latin-1')
    pdf.cell(0, 6, descrip, 0, 1)
    pdf.cell(40, 6, "RIF:", 0, 0)
    pdf.cell(0, 6, str(prov_data.get('ID3', '')), 0, 1)
    
    pdf.ln(5)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(0, 8, "Detalle de la Solicitud:", 0, 1)
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(40, 6, "ID Solicitud:", 0, 0)
    pdf.cell(0, 6, str(nc_data.get('Id', '')), 0, 1)
    pdf.cell(40, 6, "Factura Afectada:", 0, 0)
    pdf.cell(0, 6, str(nc_data.get('NumeroD', '')), 0, 1)
    pdf.cell(40, 6, "Motivo:", 0, 0)
    motivo = str(nc_data.get('Motivo', '')).encode('latin-1', 'replace').decode('latin-1')
    pdf.cell(0, 6, motivo, 0, 1)
    pdf.cell(40, 6, "Monto (Bs):", 0, 0)
    pdf.cell(0, 6, f"Bs. {float(nc_data.get('MontoBs', 0)):,.2f}", 0, 1)
    pdf.cell(40, 6, "Monto (USD):", 0, 0)
    pdf.cell(0, 6, f"$ {float(nc_data.get('MontoUsd', 0)):,.2f}", 0, 1)
    
    if nc_data.get('Observacion'):
        pdf.ln(3)
        pdf.cell(40, 6, "Observacion:", 0, 0)
        obs = str(nc_data.get('Observacion', '')).encode('latin-1', 'replace').decode('latin-1')
        pdf.multi_cell(0, 6, obs)
        
    pdf.ln(10)
    pdf.set_font('Helvetica', 'I', 9)
    body_txt = "Por favor procesar la siguiente nota de credito con los datos indicados."
    if email_template_body:
        body_txt += f"\n\nNotas Adicionales:\n- {email_template_body}"
    body = str(body_txt).encode('latin-1', 'replace').decode('latin-1')
    pdf.multi_cell(0, 5, body)
    
    return pdf.output(dest='S').encode('latin1')

def enviar_correo_solicitud_nc(destinatario: str, proveedor: dict, nc_data: dict, body_template: str) -> bool:
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    import base64
    from googleapiclient.errors import HttpError

    try:
        service = get_gmail_service()
        if not service:
            return False
            
        emails = [e.strip() for e in destinatario.split(";") if e.strip()]
        if not emails:
            return False
            
        remitente = os.getenv("SMTP_EMAIL", "")
        msg = MIMEMultipart()
        msg['From'] = remitente
        msg['To'] = ", ".join(emails)
        msg['Subject'] = f"Solicitud de Nota de Crédito - Factura {nc_data.get('NumeroD')} - {proveedor.get('Descrip')}"
        
        footer = f"\n\nNotas Adicionales:\n{body_template}" if body_template else ""
        cuerpo = f"""Estimados {proveedor.get('Descrip')} (RIF: {proveedor.get('ID3')}),

Adjunto a este correo encontrará una solicitud formal de Nota de Crédito correspondiente a la factura {nc_data.get('NumeroD')}.

Monto Solicitado: Bs. {float(nc_data.get('MontoBs', 0)):,.2f}

Quedamos a su entera disposición ante cualquier duda o comentario.

Atentamente,
El equipo de Administración.{footer}"""

        msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
        
        pdf_bytes = generar_pdf_nc_request(nc_data, proveedor, body_template)
        part_pdf = MIMEBase("application", "pdf")
        part_pdf.set_payload(pdf_bytes)
        encoders.encode_base64(part_pdf)
        part_pdf.add_header("Content-Disposition", f"attachment; filename=Solicitud_NC_{nc_data.get('NumeroD')}.pdf")
        msg.attach(part_pdf)
        
        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        service.users().messages().send(userId='me', body={'raw': raw}).execute()
        return True
    except Exception as e:
        logging.error(f"Error al enviar correo solicitud NC: {e}", exc_info=True)
        return False

@router.post("/api/procurement/credit-notes")
async def create_credit_note(payload: dict = Body(...)):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT Factor, MontoMEx FROM dbo.SACOMP WITH (NOLOCK) WHERE NumeroD = ? AND TipoCom = 'H'", (payload["NumeroD"],))
        sacomp_row = cursor.fetchone()
        tasa_orig = float(sacomp_row[0]) if sacomp_row and sacomp_row[0] else None
        mto_orig = float(sacomp_row[1]) if sacomp_row and sacomp_row[1] else None

        cursor.execute("SELECT Id FROM EnterpriseAdmin_AMC.Procurement.CreditNotesTracking WHERE CodProv = ? AND NumeroD = ? AND Motivo = ? AND Estatus IN ('PENDIENTE', 'SOLICITADA')", (payload["CodProv"], payload["NumeroD"], payload.get("Motivo", "PAGO_EXCESO")))
        if cursor.fetchone():
            raise HTTPException(status_code=400, detail="Ya existe una Nota de Crédito pendiente o solicitada para este motivo en esta factura.")

        cursor.execute("SELECT Codigo FROM EnterpriseAdmin_AMC.Procurement.MotivosAjuste WHERE Codigo = ?", (payload.get("Motivo", "PAGO_EXCESO"),))
        mot_row = cursor.fetchone()

        cursor.execute("""
            INSERT INTO EnterpriseAdmin_AMC.Procurement.CreditNotesTracking 
            (CodProv, NumeroD, Motivo, MontoBs, TasaCambio, MontoUsd, Observacion, TasaCambioOrig, MontoMExOrig, Estatus)
            OUTPUT INSERTED.Id
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            payload["CodProv"], payload["NumeroD"], 
            payload.get("Motivo", "PAGO_EXCESO"),
            float(payload["MontoBs"]),
            float(payload.get("TasaCambio", 0)),
            float(payload.get("MontoUsd", 0)),
            payload.get("Observacion", ""),
            tasa_orig, mto_orig, 'SOLICITADA' if payload.get("EnviarInmediato") else 'PENDIENTE'
        ))
        new_id = cursor.fetchone()[0]
        
        email_sent = False
        if payload.get("EnviarInmediato"):
            # Fetch Provider Email and Template
            cursor.execute("""
                SELECT p.CodProv, p.Descrip, p.ID3, c.Email 
                FROM EnterpriseAdmin_AMC.dbo.SAPROV p 
                LEFT JOIN EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones c ON p.CodProv = c.CodProv 
                WHERE p.CodProv = ?
            """, (payload["CodProv"],))
            prov_row = cursor.fetchone()
            
            cursor.execute("SELECT EmailTemplate FROM EnterpriseAdmin_AMC.Procurement.MotivosAjuste WHERE Codigo = ?", (payload.get("Motivo", "PAGO_EXCESO"),))
            tpl_row = cursor.fetchone()
            tpl = tpl_row[0] if tpl_row and tpl_row[0] else ""
            
            if prov_row and prov_row.Email:
                prov_dict = {"CodProv": prov_row.CodProv, "Descrip": prov_row.Descrip, "ID3": prov_row.ID3, "Email": prov_row.Email}
                nc_data = {
                    "Id": new_id, "NumeroD": payload["NumeroD"], "Motivo": payload.get("Motivo", "PAGO_EXCESO"),
                    "MontoBs": float(payload["MontoBs"]), "MontoUsd": float(payload.get("MontoUsd", 0)),
                    "Observacion": payload.get("Observacion", "")
                }
                email_sent = enviar_correo_solicitud_nc(prov_row.Email, prov_dict, nc_data, tpl)
                
        conn.commit()
        logging.info(f"Nota de Crédito #{new_id} creada para {payload['CodProv']} factura {payload['NumeroD']}")
        return {"message": "Nota de Crédito creada", "Id": new_id, "email_sent": email_sent, "status": 'SOLICITADA' if payload.get("EnviarInmediato") else 'PENDIENTE'}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error creating credit note: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.post("/api/procurement/credit-notes/{id_nc}/send")
async def send_credit_note_email(id_nc: int):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT cn.Id, cn.NumeroD, cn.Motivo, cn.MontoBs, cn.MontoUsd, cn.Observacion, p.CodProv, p.Descrip, p.ID3, c.Email, m.EmailTemplate
            FROM EnterpriseAdmin_AMC.Procurement.CreditNotesTracking cn
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON cn.CodProv = p.CodProv
            LEFT JOIN EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones c ON cn.CodProv = c.CodProv
            LEFT JOIN EnterpriseAdmin_AMC.Procurement.MotivosAjuste m ON cn.Motivo = m.Codigo
            WHERE cn.Id = ? AND cn.Estatus IN ('PENDIENTE', 'SOLICITADA')
        """, (id_nc,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Nota de Crédito no encontrada o no está en un estado válido para enviar.")
            
        nc_data = {
            "Id": row.Id, "NumeroD": row.NumeroD, "Motivo": row.Motivo,
            "MontoBs": row.MontoBs, "MontoUsd": row.MontoUsd, "Observacion": row.Observacion
        }
        prov_dict = {
            "CodProv": row.CodProv, "Descrip": row.Descrip, "ID3": row.ID3
        }
        email_str = row.Email
        plantilla = row.EmailTemplate or ""
        
        if not email_str:
            raise HTTPException(status_code=400, detail="El proveedor no tiene un correo configurado.")
            
        success = enviar_correo_solicitud_nc(email_str, prov_dict, nc_data, plantilla)
        if success:
            cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.CreditNotesTracking SET Estatus = 'SOLICITADA' WHERE Id = ?", (id_nc,))
            conn.commit()
            return {"message": "Correo enviado con éxito y estado actualizado a SOLICITADA."}
        else:
            raise HTTPException(status_code=500, detail="Fallo al enviar correo a través de Gmail.")
            
    except HTTPException: raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.patch("/api/procurement/credit-notes/{id_nc}")
async def update_credit_note(id_nc: int, payload: dict = Body(...)):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT Estatus, CodProv, NumeroD, MontoBs, NotaCreditoID FROM EnterpriseAdmin_AMC.Procurement.CreditNotesTracking WHERE Id = ?", (id_nc,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Nota de Crédito no encontrada")
        
        new_status = payload.get("Estatus", row[0])
        
        if new_status == "APLICADA" and row[0] != "APLICADA":
            # Apply: update status and register as abono
            nota_credito_id = payload.get("NotaCreditoID", row[4] or "")
            cursor.execute("""
                UPDATE EnterpriseAdmin_AMC.Procurement.CreditNotesTracking 
                SET Estatus = 'APLICADA', FechaEmision = GETDATE(), NotaCreditoID = ?
                WHERE Id = ?
            """, (nota_credito_id, id_nc))
            
            # Auto-register as abono
            
            cursor.execute("SELECT Factor, MontoMEx FROM dbo.SACOMP WITH (NOLOCK) WHERE NumeroD = ? AND TipoCom = 'H'", (row[2],))
            sacomp_row = cursor.fetchone()
            tasa_orig = float(sacomp_row[0]) if sacomp_row and sacomp_row[0] else None
            mto_orig = float(sacomp_row[1]) if sacomp_row and sacomp_row[1] else None

            cursor.execute("SELECT ISNULL(MAX(AbonoID), 0) FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos WITH (UPDLOCK)")
            new_abono_id = cursor.fetchone()[0] + 1

            cursor.execute("""
                INSERT INTO EnterpriseAdmin_AMC.dbo.CxP_Abonos 
                (AbonoID, NumeroD, CodProv, FechaAbono, MontoBsAbonado, TasaCambioDiaAbono, MontoUsdAbonado, AplicaIndexacion, Referencia, TipoAbono, TasaCambioOrig, MontoMExOrig, RutaComprobante, NotificarCorreo)
                VALUES (?, ?, GETDATE(), ?, 0, 0, 0, ?, 'NOTA_CREDITO', ?, ?, '', 0)
            """, (new_abono_id, row[2], row[1], float(row[3]), f"Nota Crédito #{nota_credito_id or id_nc}", tasa_orig, mto_orig))
            
        elif new_status == "ANULADA" and row[0] != "ANULADA":
            cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.CreditNotesTracking SET Estatus = 'ANULADA' WHERE Id = ?", (id_nc,))
            
            # Remove the auto-registered abono if it was applied
            if row[0] == "APLICADA":
                cursor.execute("""
                    DELETE FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos 
                    WHERE NumeroD = ? AND CodProv = ? AND TipoAbono = 'NOTA_CREDITO' 
                    AND Referencia LIKE ?
                """, (row[2], row[1], f"%{row[4] or id_nc}%"))
        else:
            # General update (e.g., editing NotaCreditoID or Observacion)
            cursor.execute("""
                UPDATE EnterpriseAdmin_AMC.Procurement.CreditNotesTracking 
                SET NotaCreditoID = COALESCE(?, NotaCreditoID), Observacion = COALESCE(?, Observacion)
                WHERE Id = ?
            """, (payload.get("NotaCreditoID"), payload.get("Observacion"), id_nc))
        
        conn.commit()
        return {"message": f"Nota de Crédito actualizada a {new_status}"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error updating credit note: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.delete("/api/procurement/credit-notes/{id_nc}")
async def delete_credit_note(id_nc: int):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT Estatus, CodProv, NumeroD, MontoBs, NotaCreditoID FROM EnterpriseAdmin_AMC.Procurement.CreditNotesTracking WHERE Id = ?", (id_nc,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Nota de Crédito no encontrada")
        
        estatus = row[0]
        # Only allow deletion for PENDIENTE or ANULADA
        # If APLICADA, we should ideally not delete or require careful handling.
        if estatus == "APLICADA":
            raise HTTPException(status_code=400, detail="No se puede eliminar una nota de crédito APLICADA. Primero anúlela.")

        # If it was APLICADA previously and we are in ANULADA, we still want to make sure abonos are gone
        # The logic in PATCH handles this, but here we just delete the record.
        cursor.execute("DELETE FROM EnterpriseAdmin_AMC.Procurement.CreditNotesTracking WHERE Id = ?", (id_nc,))
        
        conn.commit()
        return {"message": "Nota de crédito eliminada exitosamente."}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error deleting credit note: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.get("/api/procurement/settings")
async def get_settings():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT SettingKey, SettingValue FROM EnterpriseAdmin_AMC.Procurement.Settings")
        settings = {row.SettingKey: row.SettingValue for row in cursor.fetchall()}
        return settings
    except Exception as e:
        logging.error(f"Error fetching settings: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.post("/api/procurement/settings")
async def update_settings(payload: dict):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        for key, value in payload.items():
            cursor.execute("""
                UPDATE EnterpriseAdmin_AMC.Procurement.Settings 
                SET SettingValue = ?, UpdatedAt = GETDATE()
                WHERE SettingKey = ?
            """, (str(value), str(key)))
            if cursor.rowcount == 0:
                cursor.execute("""
                    INSERT INTO EnterpriseAdmin_AMC.Procurement.Settings (SettingKey, SettingValue)
                    VALUES (?, ?)
                """, (str(key), str(value)))
        conn.commit()
        return {"message": "Configuración guardada exitosamente"}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error updating settings: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.get("/api/retenciones-islr")
async def get_retenciones_islr(desde: Optional[str] = None, hasta: Optional[str] = None):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()

        query = """
            SELECT r.*, p.Descrip AS ProveedorNombre 
            FROM EnterpriseAdmin_AMC.Procurement.Retenciones_ISLR r
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON r.CodProv = p.CodProv
            WHERE 1=1
        """
        params = []
        if desde:
            query += " AND FechaRetencion >= ?"
            params.append(desde + " 00:00:00")
        if hasta:
            query += " AND FechaRetencion <= ?"
            params.append(hasta + " 23:59:59")

        query += " ORDER BY FechaRetencion DESC, Id DESC"
        cursor.execute(query, params)
        rows = cursor.fetchall()

        columns = [column[0] for column in cursor.description]
        data = [dict(zip(columns, row)) for row in rows]
        
        for r in data:
            if r.get('ProveedorNombre'):
                r['CodProv'] = f"{r['CodProv']} - {r['ProveedorNombre']}"
            for k, v in r.items():
                if hasattr(v, 'quantize'):
                    r[k] = float(v) if v is not None else 0.0
                elif hasattr(v, 'strftime'):
                    r[k] = v.strftime('%Y-%m-%d %H:%M') if v else None
                    
        return {"data": data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.post("/api/retenciones-islr")
async def crear_retencion_islr(payload: dict = Body(...)):
    """Create ISLR retention records based on Decree 1808 (supports batching)"""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        from datetime import datetime
        import uuid
        now = datetime.now()

        facturas = payload.get("facturas", [payload])
        fecha_retencion = payload.get("FechaRetencion", now.strftime('%Y-%m-%d'))
        
        # 0. Idempotency check 
        for f in facturas:
            cursor.execute("SELECT Id FROM EnterpriseAdmin_AMC.Procurement.Retenciones_ISLR WITH (NOLOCK) WHERE NumeroD = ? AND CodProv = ? AND Estado != 'ANULADO'", (f["NumeroD"], f["CodProv"]))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail=f"La factura {f['NumeroD']} ya posee una retención ISLR activa.")

        # 1. Generate sequential number for ISLR
        cursor.execute("SELECT UltimoSecuencial FROM EnterpriseAdmin_AMC.Procurement.Retenciones_ISLR_Config WITH (UPDLOCK) WHERE Id = 1")
        last_seq = cursor.fetchone()[0]
        new_seq = last_seq + 1
        
        nro_comprobante = f"{now.strftime('%Y%m')}-ISLR-{str(new_seq).zfill(4)}"
        idlote = str(uuid.uuid4())
        
        inserted_ids = []
        for f in facturas:
            # Replicating IVA insertion structure mapping explicit references
            cursor.execute("""
                INSERT INTO EnterpriseAdmin_AMC.Procurement.Retenciones_ISLR 
                (IdLote, CodProv, FechaRetencion, NumeroComprobante, NumeroD, NroCtrol, 
                 MontoTotalBs, BaseImponibleBs, AlicuotaPct, SustraendoBs, MontoRetenido, ConceptoISLR, Estado)
                OUTPUT INSERTED.Id
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'EMITIDO')
            """, (
                idlote, f["CodProv"], fecha_retencion, nro_comprobante,
                f["NumeroD"], f["NroCtrol"],
                f.get("MontoTotalBs", 0), f.get("BaseImponibleBs", 0), f.get("AlicuotaPct", 0),
                f.get("SustraendoBs", 0), f.get("MontoRetenido", 0), f.get("ConceptoISLR", "Servicios_2%")
            ))
            new_id = cursor.fetchone()[0]
            inserted_ids.append(new_id)

            monto_retenido = float(f.get("MontoRetenido", 0))

            if True:
                # Get historical exchange rate to record the abono correctly
                cursor.execute("SELECT Factor, MontoMEx FROM dbo.SACOMP WITH (NOLOCK) WHERE NumeroD = ? AND CodProv = ? AND TipoCom = 'H'", (f["NumeroD"], f["CodProv"]))
                sacomp_row = cursor.fetchone()
                tasa_orig = float(sacomp_row[0]) if sacomp_row and sacomp_row[0] else None
                mto_orig = float(sacomp_row[1]) if sacomp_row and sacomp_row[1] else None
                monto_usd_abonado = monto_retenido / tasa_orig if (tasa_orig and tasa_orig > 0) else 0

                cursor.execute("SELECT ISNULL(MAX(AbonoID), 0) FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos WITH (UPDLOCK)")
                new_abono_id = cursor.fetchone()[0] + 1

                # Auto-generate accounting Abono and Deduct Balance in SAACXP
                cursor.execute("""
                    INSERT INTO EnterpriseAdmin_AMC.dbo.CxP_Abonos 
                    (AbonoID, NumeroD, CodProv, FechaAbono, MontoBsAbonado, TasaCambioDiaAbono, MontoUsdAbonado, AplicaIndexacion, Referencia, TipoAbono, TasaCambioOrig, MontoMExOrig, RutaComprobante, NotificarCorreo)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    new_abono_id, f["NumeroD"], f["CodProv"], fecha_retencion,
                    monto_retenido, tasa_orig or 0, monto_usd_abonado, 0,
                    f"Retención ISLR Comp. {nro_comprobante}", 'RETENCION_ISLR',
                    tasa_orig, mto_orig, '', 0
                ))
        
        # 3. Update Config Sequential
        cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.Retenciones_ISLR_Config SET UltimoSecuencial = ? WHERE Id = 1", (new_seq,))
        
        conn.commit()
        logging.info(f"Retención ISLR {nro_comprobante} creada con {len(facturas)} factura(s)")
        return {"message": "Retención ISLR creada", "NumeroComprobante": nro_comprobante, "Ids": inserted_ids}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error creating ISLR retencion: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@router.patch("/api/retenciones-islr/{id_ret}")
async def anular_retencion_islr(id_ret: int):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT Estado, NumeroComprobante, NumeroD, CodProv FROM EnterpriseAdmin_AMC.Procurement.Retenciones_ISLR WHERE Id = ?", (id_ret,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Retención ISLR no encontrada")
        if row[0] == 'ENTERADO':
            raise HTTPException(status_code=400, detail="No se puede anular una retención ENTERADA (declarada ante SENIAT)")
            
        cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.Retenciones_ISLR SET Estado = 'ANULADO' WHERE Id = ?", (id_ret,))
        
        # Delete the auto-registered abono
        nro_comprobante = row[1]
        cursor.execute("""
            DELETE FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos 
            WHERE NumeroD = ? AND CodProv = ? AND TipoAbono = 'RETENCION_ISLR' 
            AND Referencia LIKE ?
        """, (row[2], row[3], f"%{nro_comprobante}%"))
        
        conn.commit()
        return {"message": "Retención ISLR anulada y abono correspondiente eliminado"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

# --- PDF Generation for ISLR ---
def generar_pdf_islr(config: dict, retenciones: list) -> bytes:
    from fpdf import FPDF
    from datetime import datetime

    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    # Colors
    header_bg = (0, 102, 51)   # Dark green for ISLR
    header_fg = (255, 255, 255)
    row_alt = (230, 250, 240)
    
    # --- Header ---
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_fill_color(*header_bg)
    pdf.set_text_color(*header_fg)
    pdf.cell(0, 10, 'COMPROBANTE DE RETENCIÓN DE ISLR', 0, 1, 'C', fill=True)
    pdf.ln(3)
    
    # --- Agente info ---
    nro_comprobante = retenciones[0].get("NumeroComprobante", "N/A") if retenciones else "N/A"
    fecha_ret = str(retenciones[0].get("FechaRetencion", ""))[:10] if retenciones else "N/A"

    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "Agente de Retención:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, str(config.get("RazonSocial_Agente", "")), 0, 1)

    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "RIF:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, str(config.get("RIF_Agente", "")), 0, 1)

    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "Dirección Fiscal:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    pdf.multi_cell(0, 6, str(config.get("DireccionFiscal_Agente", "")), 0, 'L')
    pdf.set_x(15)
    
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "Nro. Comprobante:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, str(nro_comprobante), 0, 1)

    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "Fecha de Retención:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, str(fecha_ret), 0, 1)
    pdf.ln(3)
    
    # --- Sujeto Retenido ---
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "Sujeto Retenido:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    prov_name = retenciones[0].get("ProveedorNombre") or retenciones[0].get("CodProv", "")
    pdf.cell(0, 6, str(prov_name), 0, 1)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "RIF Proveedor:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, str(retenciones[0].get("CodProv", "")), 0, 1)
    pdf.ln(5)
    
    # --- Table Header ---
    cols = [
        ("Factura", 25), ("Nro Control", 25), 
        ("Monto Total", 25), ("Base Imponible", 25), 
        ("Ret. %", 15), ("Sustraendo", 25), ("Monto Retenido", 25)
    ]
    
    pdf.set_font('Helvetica', 'B', 7)
    pdf.set_fill_color(*header_bg)
    pdf.set_text_color(*header_fg)
    for name, width in cols:
        pdf.cell(width, 7, name, 1, 0, 'C', fill=True)
    pdf.ln()
    
    # --- Table Rows ---
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 7)
    
    total_monto = 0
    total_base = 0
    total_sustraendo = 0
    total_retenido = 0
    
    for i, r in enumerate(retenciones):
        if i % 2 == 1:
            pdf.set_fill_color(*row_alt)
            fill = True
        else:
            fill = False
            
        monto = float(r.get("MontoTotalBs") or r.get("MontoTotal", 0))
        base = float(r.get("BaseImponibleBs") or r.get("BaseImponible", 0))
        alicuota = float(r.get("AlicuotaPct", 0))
        sustraendo = float(r.get("SustraendoBs", 0))
        retenido = float(r.get("MontoRetenido", 0))
        
        total_monto += monto
        total_base += base
        total_sustraendo += sustraendo
        total_retenido += retenido
        
        row_data = [
            str(r.get("NumeroD", "")),
            str(r.get("NroCtrol", "")),
            f"{monto:,.2f}",
            f"{base:,.2f}",
            f"{alicuota:.2f}%",
            f"{sustraendo:,.2f}",
            f"{retenido:,.2f}"
        ]
        for j, (name, width) in enumerate(cols):
            align = 'R' if j >= 2 else 'L'
            pdf.cell(width, 6, row_data[j], 1, 0, align, fill=fill)
        pdf.ln()
    
    # --- Totals ---
    pdf.set_font('Helvetica', 'B', 7)
    pdf.set_fill_color(220, 220, 220)
    
    prep_width = cols[0][1] + cols[1][1]
    pdf.cell(prep_width, 7, 'TOTALES:', 1, 0, 'R', fill=True)
    pdf.cell(cols[2][1], 7, f"{total_monto:,.2f}", 1, 0, 'R', fill=True)
    pdf.cell(cols[3][1], 7, f"{total_base:,.2f}", 1, 0, 'R', fill=True) 
    pdf.cell(cols[4][1], 7, "", 1, 0, 'C', fill=True)
    pdf.cell(cols[5][1], 7, f"{total_sustraendo:,.2f}", 1, 0, 'R', fill=True)
    pdf.cell(cols[6][1], 7, f"{total_retenido:,.2f}", 1, 0, 'R', fill=True)
    pdf.ln(12)
    
    # --- Footer ---
    pdf.set_font('Helvetica', '', 8)
    pdf.cell(0, 5, f"Documento generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", 0, 1, 'C')
    
    out = pdf.output(dest='S')
    return out.encode('latin-1') if isinstance(out, str) else bytes(out)

if __name__ == "__main__":
    import uvicorn # type: ignore
    uvicorn.run(app, host="0.0.0.0", port=8080)
